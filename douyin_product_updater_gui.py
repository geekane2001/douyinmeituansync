import requests
import json
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import traceback
import pandas as pd
from openai import OpenAI
import hashlib
import re
import base64
from PIL import Image
import boto3
from botocore.client import Config
from io import BytesIO
import datetime
import logging

# --- 全局日志配置 ---
log_file_path = 'update_products_log.txt'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path, encoding='utf-8'),
        # logging.StreamHandler() # 如果也想在控制台看到日志，可以取消这行注释
    ]
)

# --- 1. 配置信息 ---

# Cloudflare R2 配置
CLOUDFLARE_ACCOUNT_ID = "67a7569d0cd89aafb7499f3cf3bc9f73"
CLOUDFLARE_R2_ACCESS_KEY_ID = "6684b2a5b8f947ba4f6f3ba943d22439"
CLOUDFLARE_R2_SECRET_ACCESS_KEY = "bd3dce5ac2df30ae34377c9ca5af26fd845abe5fa6ea179ec6810552856ca27f"
R2_BUCKET_NAME = "0926taocantoutu"
R2_ENDPOINT_URL = f"https://{CLOUDFLARE_ACCOUNT_ID}.r2.cloudflarestorage.com"
R2_PUBLIC_URL_PREFIX = "https://pub-c92931353257460eb0beccbf59ef2ad0.r2.dev"

# ModelScope LLM (DeepSeek) 配置
MS_BASE_URL = 'https://api-inference.modelscope.cn/v1'
MS_API_KEY = 'ms-871a8344-b18d-4fb5-b96e-d4123fbbb0f0'
LLM_MODEL_ID = 'deepseek-ai/DeepSeek-V3.2-Exp'
VISION_MODEL_IDS = [
    'Qwen/Qwen3-VL-8B-Instruct',
    'Qwen/Qwen3-VL-235B-A22B-Instruct',
    'Qwen/Qwen3-VL-30B-A3B-Instruct'
]

# LLM Client with timeout
try:
    llm_client = OpenAI(base_url=MS_BASE_URL, api_key=MS_API_KEY, timeout=30.0)
except Exception as e:
    llm_client = None
    print(f"初始化LLM客户端失败: {e}")

# 抖音开放平台密钥
CLIENT_KEY = "awbeykzyos7kbidv"
CLIENT_SECRET = "4575440b156ecbe144284e4f69d284a2"
DOUYIN_ACCOUNT_ID = "7241078611527075855"

# 抖音网页端配置（用于重创模式）
DOUYIN_WEB_CSRF_TOKEN = "000100000001ae8a406b9344d0cc4e30ceaf542c505dbbabca5a3842c450a93e0787a4d2f8991880c8ea9d2d1372"
DOUYIN_ROOT_LIFE_ACCOUNT_ID = "7241078611527075855"  # 根账号ID（与DOUYIN_ACCOUNT_ID相同）

# 从文件读取Cookie
def load_cookie_from_file():
    """从cookie.txt文件读取Cookie"""
    cookie_file = os.path.join(os.path.dirname(__file__), 'cookie.txt')
    try:
        with open(cookie_file, 'r', encoding='utf-8') as f:
            cookie = f.read().strip()
            if cookie:
                logging.info(f"成功从 {cookie_file} 加载Cookie")
                return cookie
    except FileNotFoundError:
        logging.warning(f"未找到Cookie文件: {cookie_file}")
    except Exception as e:
        logging.error(f"读取Cookie文件失败: {e}")
    return ""

DOUYIN_WEB_COOKIE = load_cookie_from_file()

# 飞书多维表格配置
FEISHU_APP_ID = "cli_a6672cae343ad00e"
FEISHU_APP_SECRET = "0J4SpfBMeIxJEOXDJMNbofMipRgwkMpV"
FEISHU_APP_TOKEN = "MslRbdwPca7P6qsqbqgcvpBGnRh"
FEISHU_TABLE_ID = "tbluVbrXLRUmfouv"

# --- 2. API 和 URL 地址 ---
DOUYIN_TOKEN_URL = "https://open.douyin.com/oauth/client_token/"
DOUYIN_PRODUCT_QUERY_URL = "https://open.douyin.com/goodlife/v1/goods/product/online/query/"
DOUYIN_PRODUCT_GET_URL = "https://open.douyin.com/goodlife/v1/goods/product/online/get/"
DOUYIN_PRODUCT_SAVE_URL = "https://open.douyin.com/goodlife/v1/goods/product/save/"
DOUYIN_PRODUCT_OPERATE_URL = "https://open.douyin.com/goodlife/v1/goods/product/operate/"
FEISHU_TENANT_ACCESS_TOKEN_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
FEISHU_BITABLE_RECORDS_SEARCH_URL = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{FEISHU_APP_TOKEN}/tables/{FEISHU_TABLE_ID}/records/search"

# --- 3. 后端业务逻辑 ---

def get_feishu_tenant_access_token(log_func):
    log_func("--- 正在获取飞书 Access Token ---")
    try:
        response = requests.post(FEISHU_TENANT_ACCESS_TOKEN_URL, json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET}, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get("code") == 0:
            log_func("[Success] 飞书 Token 获取成功!")
            return data.get("tenant_access_token")
    except Exception as e:
        log_func(f"[Error] 获取飞书 Token 失败: {e}")
    return None

def get_feishu_bitable_records(feishu_token, log_func):
    log_func("--- 正在从飞书获取门店列表 ---")
    headers = {"Authorization": f"Bearer {feishu_token}"}
    all_records = {}
    page_token = ""
    while True:
        params = {"page_size": 500, "page_token": page_token}
        try:
            response = requests.post(FEISHU_BITABLE_RECORDS_SEARCH_URL, headers=headers, params=params, json={"field_names": ["门店名称", "门店ID"]}, timeout=15)
            response.raise_for_status()
            data = response.json()
            if data.get("code") != 0:
                log_func(f"[Error] 查询飞书记录失败: {data.get('msg')}")
                return {}
            items = data.get("data", {}).get("items", [])
            for item in items:
                fields = item.get("fields", {})
                store_name = fields.get("门店名称", [{}])[0].get('text')
                store_id = fields.get("门店ID", [{}])[0].get('text')
                if store_name and store_id: all_records[store_name] = store_id
            page_token = data.get("data", {}).get("page_token")
            if not data.get("data", {}).get("has_more", False): break
        except Exception as e:
            log_func(f"[Error] 查询飞书记录时发生错误: {e}")
            return {}
    log_func(f"[Success] 成功从飞书获取到 {len(all_records)} 个门店。")
    return all_records

def get_douyin_access_token(log_func):
    log_func("--- 正在获取抖音 Access Token ---")
    try:
        response = requests.post(DOUYIN_TOKEN_URL, json={"grant_type": "client_credential", "client_key": CLIENT_KEY, "client_secret": CLIENT_SECRET}, timeout=10)
        response.raise_for_status()
        data = response.json().get("data", {})
        if data.get("error_code") == 0 and data.get("access_token"):
            log_func("[Success] 抖音 Token 获取成功!")
            return data["access_token"]
    except Exception as e:
        log_func(f"[Error] 获取抖音 Token 失败: {e}")
    return None

def get_douyin_products_by_store(access_token, poi_id, log_func):
    log_func(f"--- 正在使用 POI ID: {poi_id} 查询抖音商品列表 ---")
    headers = {"Content-Type": "application/json", "access-token": access_token}
    params = {"account_id": str(DOUYIN_ACCOUNT_ID), "poi_ids": f'[{poi_id}]', "count": 50, "status": 1}
    try:
        response = requests.get(DOUYIN_PRODUCT_QUERY_URL, headers=headers, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        if data.get("data", {}).get("error_code") != 0: return []
        product_list = data.get("data", {}).get("products", [])
        detailed_products = []
        for p in product_list:
            product_info = p.get("product", {})
            sku_info = p.get("sku", {})
            if product_info and sku_info:
                detailed_products.append({
                    "id": product_info.get('product_id'),
                    "name": product_info.get('product_name'),
                    "price": f"{sku_info.get('actual_amount', 0) / 100:.2f}",
                    "origin_price": f"{sku_info.get('origin_amount', 0) / 100:.2f}"
                })
        log_func(f"[Success] 查询到 {len(detailed_products)} 个在线商品。")
        return detailed_products
    except Exception as e:
        log_func(f"[Error] 查询抖音商品时发生错误: {e}")
    return []

def get_douyin_product_details(access_token, product_id, log_func):
    log_func(f"--- 正在获取商品 '{product_id}' 的详细信息 ---")
    headers = {"Content-Type": "application/json", "access-token": access_token}
    params = {"account_id": DOUYIN_ACCOUNT_ID, "product_ids": [product_id]}
    try:
        response = requests.get(DOUYIN_PRODUCT_GET_URL, headers=headers, params=params, timeout=10)
        response.raise_for_status()
        response_data = response.json()
        if response_data.get("data", {}).get("error_code") == 0:
            product_details = response_data.get("data", {}).get("product_onlines", [])
            if product_details:
                log_func(f"[Success] 商品 '{product_id}' 详情获取成功。")
                return product_details[0]
    except Exception as e:
        log_func(f"[Error] 获取商品 '{product_id}' 详情时发生意外错误: {e}")
    return None

def operate_douyin_product(access_token, product_id, log_func, offline=True):
    op_type = 2 if offline else 1
    action_text = "下架" if offline else "上架"
    log_func(f"========== 开始 {action_text} 商品 ID: {product_id} ==========")
    headers = {"Content-Type": "application/json", "access-token": access_token}
    payload = { "account_id": DOUYIN_ACCOUNT_ID, "product_id": product_id, "op_type": op_type }
    try:
        response = requests.post(DOUYIN_PRODUCT_OPERATE_URL, headers=headers, json=payload, timeout=20)
        response.raise_for_status()
        response_data = response.json()
        if response_data.get('data', {}).get('error_code') == 0:
            log_func(f"[SUCCESS] 商品 {product_id} {action_text}成功!")
            return True, ""
        else:
            reason = response_data.get('data', {}).get('description', 'API返回未知错误')
            log_func(f"[FAILURE] 商品 {product_id} {action_text}失败: {reason}")
            return False, reason
    except Exception as e:
        log_func(f"商品 {product_id} {action_text}时发生意外错误: {e}")
        return False, str(e)

# --- 美团同步相关函数 ---
def process_store_name_for_meituan(store_name, log_func):
    """处理店名用于美团搜索（移除'竞潮玩'）"""
    cleaned_name = store_name.replace("竞潮玩", "").strip()
    log_func(f"处理店名: '{store_name}' -> '{cleaned_name}'")
    return cleaned_name

def get_meituan_packages(store_name, city, log_func):
    """获取美团套餐信息"""
    from bs4 import BeautifulSoup
    import time
    import re
    
    url = f"https://i.meituan.com/s/{city}-{store_name}"
    log_func(f"--- 正在请求美团URL: {url} ---")
    
    # 美团Cookie模板
    current_timestamp_ms = int(time.time() * 1000)
    base_cookie = (
        f"__mta=176011805.1756208359328.{current_timestamp_ms-5000}.{current_timestamp_ms}.30; "
        "iuuid=BB0697D3630DED2F82ADB96105EC195EB173E4FFD90723B66428C9829840A7AA; "
        "_lxsdk_cuid=199985447cbc8-0682d08c20b0dd-4c657b58-1fa400-199985447cbc8; "
        "_lxsdk=BB0697D3630DED2F82ADB96105EC195EB173E4FFD90723B66428C9829840A7AA; "
        "uuid=15efb5c50ad74159b2cd.1759197288.1.0.0; "
        "webp=1; "
        "_hc.v=17500b8b-5eb7-fe17-d4e8-ffc997c5aeda.1762150863; "
        "token=AgE9Jw3iB0xS0K4Dvg5h7_SFKSHEFNG3l3ns5orAsvwiPjQSJEe4ONv8nXX8acfUcNWMJhCiWxrpXgAAAAD1LgAAJsMNtP1gv1zb-teZ_5_kWSrKWuemK26NdJ5W9PXcqThYpPwqaiFNJIXoQXyCW92I; "
        "userId=4976202507; "
        f"latlng=30.602421,104.09746,{current_timestamp_ms}; "
        f"_lxsdk_s=199869550cd-0fd-14f-716%7C%7C46"
    )
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Host': 'i.meituan.com',
        'Cookie': base_cookie
    }
    
    proxies = {'http': 'http://127.0.0.1:10808', 'https': 'http://127.0.0.1:10808'}
    
    try:
        response = requests.get(url, headers=headers, proxies=proxies, timeout=15)
        response.raise_for_status()
        response.encoding = 'utf-8'
        
        log_func(f"请求成功! 状态码: {response.status_code}")
        
        if "访问异常" in response.text:
            log_func("[Error] 美团页面访问异常，可能需要更新Cookie")
            log_func(f"[Debug] 页面部分内容: {response.text[:500]}")
            return []
        
        # 检查是否包含店名
        if store_name not in response.text:
            log_func(f"[Warning] 页面中未找到店名'{store_name}'，可能搜索失败")
            log_func(f"[Debug] 页面部分内容: {response.text[:500]}")
        
        soup = BeautifulSoup(response.text, 'lxml')
        
        # 提取店名用于确认
        shop_name_tag = soup.find('span', class_='poiname')
        shop_name = shop_name_tag.text.strip() if shop_name_tag else "未知店名"
        log_func(f"--- 成功提取到【{shop_name}】的美团套餐信息 ---")
        
        deal_items = soup.select('dl.bd-deal-list dd a.react')
        
        if not deal_items:
            log_func("[Warning] 未找到美团套餐信息 (CSS选择器: dl.bd-deal-list dd a.react)")
            log_func(f"[Debug] 页面内容前1000字符: {response.text[:1000]}")
            # 保存完整HTML用于调试
            debug_file = f"meituan_debug_{int(time.time())}.html"
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(response.text)
            log_func(f"[Debug] 完整HTML已保存到: {debug_file}")
            return []
        
        log_func(f"找到 {len(deal_items)} 个套餐项")
        
        packages = []
        for idx, item in enumerate(deal_items):
            log_func(f"\n========== 开始解析套餐 {idx+1} ==========")
            
            # 打印该套餐项的完整HTML结构（前500字符）
            item_html = str(item)[:500]
            log_func(f"[HTML结构] 套餐{idx+1}的HTML前500字符:\n{item_html}")
            
            # 提取标题
            title_tag = item.find('div', class_='title')
            log_func(f"[步骤1-标题] title_tag是否找到: {title_tag is not None}")
            if title_tag:
                log_func(f"[步骤1-标题] title_tag内容: {title_tag}")
                title = title_tag.text.strip()
                log_func(f"[步骤1-标题] 提取的标题: '{title}'")
            else:
                title = "无标题"
                log_func(f"[步骤1-标题] 未找到标题，使用默认值: '{title}'")
            
            # 提取现价 - 详细分析
            log_func(f"[步骤2-现价] 开始查找现价...")
            price_tag = item.find('span', class_='strong')
            log_func(f"[步骤2-现价] 使用选择器'span.strong'查找结果: {price_tag is not None}")
            
            if price_tag:
                log_func(f"[步骤2-现价] 找到price_tag: {price_tag}")
                price_str = price_tag.text.strip()
                log_func(f"[步骤2-现价] price_tag.text.strip() = '{price_str}'")
            else:
                # 尝试其他选择器
                log_func(f"[步骤2-现价] 'span.strong'未找到，尝试其他选择器...")
                
                # 尝试所有span标签
                all_spans = item.find_all('span')
                log_func(f"[步骤2-现价] 该套餐项中共有 {len(all_spans)} 个span标签")
                for span_idx, span in enumerate(all_spans):
                    log_func(f"[步骤2-现价]   span[{span_idx}]: class={span.get('class')}, text='{span.text.strip()}'")
                
                price_str = "无价格"
                log_func(f"[步骤2-现价] 所有选择器均未找到，使用默认值: '{price_str}'")
            
            # 提取原价 - 详细分析
            log_func(f"[步骤3-原价] 开始查找原价...")
            original_price_tag = item.find('del')
            log_func(f"[步骤3-原价] 使用选择器'del'查找结果: {original_price_tag is not None}")
            
            if original_price_tag:
                log_func(f"[步骤3-原价] 找到original_price_tag: {original_price_tag}")
                original_price_str = original_price_tag.text.strip()
                log_func(f"[步骤3-原价] original_price_tag.text.strip() = '{original_price_str}'")
            else:
                log_func(f"[步骤3-原价] 'del'标签未找到，尝试其他选择器...")
                
                # 尝试s标签
                s_tag = item.find('s')
                log_func(f"[步骤3-原价] 使用选择器's'查找结果: {s_tag is not None}")
                if s_tag:
                    original_price_str = s_tag.text.strip()
                    log_func(f"[步骤3-原价] 从's'标签提取: '{original_price_str}'")
                else:
                    original_price_str = ""
                    log_func(f"[步骤3-原价] 所有选择器均未找到，使用空字符串")
            
            log_func(f"[步骤4-汇总] 原始提取结果: title='{title}', price_str='{price_str}', original_price_str='{original_price_str}'")
            
            # 转换为数字
            log_func(f"[步骤5-转换] 开始将字符串转换为数字...")
            try:
                # 清理价格字符串
                log_func(f"[步骤5-转换] 清理现价字符串: '{price_str}'")
                price_clean = re.sub(r'[^\d.]', '', price_str)
                log_func(f"[步骤5-转换] 清理后的现价: '{price_clean}'")
                
                if price_clean:
                    price = float(price_clean)
                    log_func(f"[步骤5-转换] 现价转换成功: {price}")
                else:
                    price = 0.0
                    log_func(f"[步骤5-转换] 清理后为空，现价设为: {price}")
                
                if original_price_str:
                    log_func(f"[步骤5-转换] 清理原价字符串: '{original_price_str}'")
                    original_price_clean = re.sub(r'[^\d.]', '', original_price_str)
                    log_func(f"[步骤5-转换] 清理后的原价: '{original_price_clean}'")
                    
                    if original_price_clean:
                        original_price = float(original_price_clean)
                        log_func(f"[步骤5-转换] 原价转换成功: {original_price}")
                    else:
                        original_price = price
                        log_func(f"[步骤5-转换] 清理后为空，原价使用现价: {original_price}")
                else:
                    original_price = price
                    log_func(f"[步骤5-转换] 原价字符串为空，使用现价: {original_price}")
                
                log_func(f"[步骤5-转换] ✅ 转换完成: price={price}, original_price={original_price}")
                
            except (ValueError, AttributeError) as e:
                log_func(f"[步骤5-转换] ❌ 转换失败: {e}")
                log_func(f"[步骤5-转换] 失败时的原始值: price_str='{price_str}', original_price_str='{original_price_str}'")
                price = 0.0
                original_price = 0.0
                log_func(f"[步骤5-转换] 设置默认值: price={price}, original_price={original_price}")
            
            packages.append({
                "title": title,
                "price": price,
                "original_price": original_price
            })
            log_func(f"[步骤6-完成] ✅ 套餐{idx+1}解析完成: {title} | 现价: {price}元 | 原价: {original_price}元")
            log_func(f"========== 套餐 {idx+1} 解析结束 ==========\n")
        
        log_func(f"成功获取 {len(packages)} 个美团套餐")
        return packages
        
    except Exception as e:
        log_func(f"[Error] 获取美团套餐失败: {e}")
        import traceback
        log_func(f"[Debug] 详细错误: {traceback.format_exc()}")
        return []

def match_packages_smart(douyin_packages, meituan_packages, log_func):
    """智能匹配抖音和美团套餐（基于现价和原价的相似度）"""
    log_func("\n" + "="*80)
    log_func("开始智能匹配抖音和美团套餐")
    log_func("="*80)
    
    # 打印输入数据概览
    log_func(f"\n[输入数据] 抖音套餐数量: {len(douyin_packages)}")
    log_func(f"[输入数据] 美团套餐数量: {len(meituan_packages)}")
    
    # 详细打印抖音套餐信息
    log_func("\n--- 抖音套餐详细信息 ---")
    for idx, dy_pkg in enumerate(douyin_packages):
        log_func(f"抖音套餐[{idx+1}]:")
        log_func(f"  ID: {dy_pkg.get('id')}")
        log_func(f"  名称: {dy_pkg.get('name')}")
        log_func(f"  现价(price): {dy_pkg.get('price')} (类型: {type(dy_pkg.get('price'))})")
        log_func(f"  原价(origin_price): {dy_pkg.get('origin_price')} (类型: {type(dy_pkg.get('origin_price'))})")
    
    # 详细打印美团套餐信息
    log_func("\n--- 美团套餐详细信息 ---")
    for idx, mt_pkg in enumerate(meituan_packages):
        log_func(f"美团套餐[{idx+1}]:")
        log_func(f"  标题: {mt_pkg.get('title')}")
        log_func(f"  现价(price): {mt_pkg.get('price')} (类型: {type(mt_pkg.get('price'))})")
        log_func(f"  原价(original_price): {mt_pkg.get('original_price')} (类型: {type(mt_pkg.get('original_price'))})")
    
    matches = []  # 匹配成功的
    meituan_only = []  # 美团独有的（需要新建）
    douyin_only = []  # 抖音独有的（需要下架，除了特殊套餐）
    
    # 特殊套餐列表（不下架）
    special_packages = ["【新老会员】28得30网费", "28得30网费"]
    
    # 为每个美团套餐寻找匹配的抖音套餐
    matched_douyin_ids = set()
    
    log_func("\n" + "-"*80)
    log_func("【第一轮匹配】优先匹配价格完全相同的套餐")
    log_func("-"*80)
    
    # 第一轮：只匹配价格完全相同的套餐（现价和原价差异都<0.01元）
    for mt_idx, mt_pkg in enumerate(meituan_packages):
        mt_price = mt_pkg['price']
        mt_orig_price = mt_pkg['original_price']
        
        log_func(f"\n>>> 第一轮 - 美团套餐 [{mt_idx+1}/{len(meituan_packages)}]: {mt_pkg['title']} <<<")
        log_func(f"  现价: {mt_price}, 原价: {mt_orig_price}")
        
        best_match = None
        best_score = 0
        
        for dy_idx, dy_pkg in enumerate(douyin_packages):
            if dy_pkg['id'] in matched_douyin_ids:
                continue
            
            try:
                dy_price = float(dy_pkg['price'])
                dy_orig_price = float(dy_pkg['origin_price'])
            except Exception as e:
                log_func(f"  ❌ 抖音套餐价格转换失败: {dy_pkg['name']}, 错误: {e}")
                continue
            
            # 计算价格差异
            price_diff = abs(mt_price - dy_price)
            orig_price_diff = abs(mt_orig_price - dy_orig_price)
            
            # 第一轮只匹配价格完全相同的
            if price_diff < 0.01 and orig_price_diff < 0.01:
                score = 100  # 完全匹配得满分
                log_func(f"  ✓ 找到完全匹配: {dy_pkg['name']} (现价:{dy_price}, 原价:{dy_orig_price})")
                
                if score > best_score:
                    best_score = score
                    best_match = dy_pkg
        
        if best_match:
            matches.append({
                "douyin": best_match,
                "meituan": mt_pkg,
                "action": "keep"  # 价格完全相同，保持原样
            })
            matched_douyin_ids.add(best_match['id'])
            log_func(f"  ✅ 第一轮匹配成功 - 价格完全相同，保持原样")
            log_func(f"     抖音: {best_match['name']}")
            log_func(f"     美团: {mt_pkg['title']}")
    
    log_func("\n" + "-"*80)
    log_func(f"【第一轮匹配完成】成功匹配 {len(matches)} 个完全相同的套餐")
    log_func("-"*80)
    
    log_func("\n" + "-"*80)
    log_func("【第二轮匹配】匹配价格相似的套餐")
    log_func("-"*80)
    
    # 第二轮：匹配价格相似的套餐（现价差异≤2元且原价差异≤30元）
    for mt_idx, mt_pkg in enumerate(meituan_packages):
        # 检查是否已在第一轮匹配
        already_matched = any(m['meituan']['title'] == mt_pkg['title'] for m in matches)
        if already_matched:
            log_func(f"\n>>> 第二轮 - 美团套餐 [{mt_idx+1}]: {mt_pkg['title']} - 已在第一轮匹配，跳过 <<<")
            continue
        
        mt_price = mt_pkg['price']
        mt_orig_price = mt_pkg['original_price']
        
        log_func(f"\n>>> 第二轮 - 美团套餐 [{mt_idx+1}/{len(meituan_packages)}]: {mt_pkg['title']} <<<")
        log_func(f"  现价: {mt_price}, 原价: {mt_orig_price}")
        
        best_match = None
        best_score = 0
        
        for dy_idx, dy_pkg in enumerate(douyin_packages):
            if dy_pkg['id'] in matched_douyin_ids:
                continue
            
            log_func(f"  比对抖音套餐: {dy_pkg['name']}")
            
            try:
                dy_price = float(dy_pkg['price'])
                dy_orig_price = float(dy_pkg['origin_price'])
                log_func(f"    抖音现价: {dy_price}, 抖音原价: {dy_orig_price}")
            except Exception as e:
                log_func(f"    ❌ 价格转换失败: {e}")
                continue
            
            # 计算价格差异
            price_diff = abs(mt_price - dy_price)
            orig_price_diff = abs(mt_orig_price - dy_orig_price)
            
            log_func(f"    现价差异: {price_diff:.2f}元, 原价差异: {orig_price_diff:.2f}元")
            
            # 第二轮匹配逻辑：价格相似
            is_match = False
            match_reason = ""
            
            if price_diff <= 0.5:
                # 现价几乎相同
                is_match = True
                match_reason = "现价几乎相同"
                score = 100 - (price_diff * 20) - (orig_price_diff * 0.5)
            elif price_diff <= 2.0 and orig_price_diff <= 30.0:
                # 现价和原价都在容忍范围内
                is_match = True
                match_reason = "现价和原价都相似"
                score = 100 - (price_diff * 10) - (orig_price_diff * 1)
            
            if is_match:
                log_func(f"    ✓ 符合匹配条件！原因: {match_reason}, 分数: {score:.1f}")
                
                if score > best_score:
                    best_score = score
                    best_match = dy_pkg
                    log_func(f"    ★ 当前最佳匹配！")
            else:
                log_func(f"    ✗ 不符合匹配条件")
        
        if best_match:
            # 判断是否需要更新价格
            dy_price = float(best_match['price'])
            dy_orig_price = float(best_match['origin_price'])
            price_diff = abs(mt_price - dy_price)
            orig_price_diff = abs(mt_orig_price - dy_orig_price)
            
            # 判断操作类型
            if price_diff <= 2.0 and orig_price_diff < 0.01:
                action = "keep"  # 现价差异小且原价相同，保持原样
                log_func(f"  ✅ 第二轮匹配成功 - 现价差异小（{price_diff:.1f}元）且原价相同，保持原样")
            else:
                action = "update"  # 需要更新价格
                log_func(f"  ✅ 第二轮匹配成功 - 价格差异较大，需要更新")
            
            matches.append({
                "douyin": best_match,
                "meituan": mt_pkg,
                "action": action
            })
            matched_douyin_ids.add(best_match['id'])
            log_func(f"     抖音: {best_match['name']} (现价:{dy_price}, 原价:{dy_orig_price})")
            log_func(f"     美团: {mt_pkg['title']} (现价:{mt_price}, 原价:{mt_orig_price})")
            log_func(f"     匹配分数: {best_score:.1f}")
            log_func(f"     操作: {action}")
        else:
            meituan_only.append(mt_pkg)
            log_func(f"  ❌ 未找到匹配的抖音套餐")
            log_func(f"     美团套餐: {mt_pkg['title']} (现价:{mt_price}, 原价:{mt_orig_price})")
            log_func(f"     → 标记为需要新建")
    
    log_func("\n" + "-"*80)
    log_func(f"【第二轮匹配完成】")
    log_func("-"*80)
    
    # 找出抖音独有的套餐
    log_func("\n" + "-"*80)
    log_func("检查抖音独有套餐")
    log_func("-"*80)
    
    for dy_pkg in douyin_packages:
        if dy_pkg['id'] not in matched_douyin_ids:
            if dy_pkg['name'] not in special_packages:
                douyin_only.append(dy_pkg)
                log_func(f"⚠️ 抖音独有套餐（需下架）: {dy_pkg['name']}")
            else:
                log_func(f"🔒 特殊套餐（保留不下架）: {dy_pkg['name']}")
    
    log_func("\n" + "="*80)
    log_func("匹配结果汇总")
    log_func("="*80)
    
    # 统计不同操作类型的数量
    keep_count = sum(1 for m in matches if m["action"] == "keep")
    update_count = sum(1 for m in matches if m["action"] == "update")
    
    log_func(f"✅ 成功匹配: {len(matches)} 个")
    log_func(f"   - 保持原样: {keep_count} 个")
    log_func(f"   - 需要更新: {update_count} 个")
    log_func(f"➕ 需要新建: {len(meituan_only)} 个")
    log_func(f"➖ 需要下架: {len(douyin_only)} 个")
    log_func("="*80 + "\n")
    
    return {
        "matches": matches,
        "meituan_only": meituan_only,
        "douyin_only": douyin_only
    }

# --- 网页端API创建商品（用于重创模式）---
def _get_product_template_web(session, product_id, root_life_account_id, log_func):
    """从网页端获取商品模板"""
    log_func(f"--- [网页端] 正在获取商品模板 (ID: {product_id})... ---")
    url = "https://life.douyin.com/life/tobias/product/get/"
    params = {
        'product_type': '1',
        'category_id': '4007001',
        'scene': '2',
        'product_id': product_id,
        'list_tab': '9',
        'source': '1',
        'is_lite_req': 'false',
        'root_life_account_id': root_life_account_id
    }
    
    try:
        response = session.get(url, params=params, timeout=20)
        response.raise_for_status()
        result = response.json()
        
        if result.get('status_code') == 0 and result.get('product_detail'):
            log_func("✅ 成功获取商品模板！")
            return result['product_detail'], None
        else:
            return None, f"获取模板失败: {result.get('status_msg', '未知错误')}"
    except requests.exceptions.RequestException as e:
        return None, f"获取模板请求失败: {e}"

def _build_web_product_payload_from_template(product_detail_template, new_data, log_func):
    """基于模板构建网页端商品创建payload（复用图片）"""
    log_func("--- [网页端] 正在基于模板构建商品负载... ---")
    
    # 移除不需要的字段
    product_detail_template.pop('product_permission_list', None)
    
    if 'product' not in product_detail_template:
        return None
    
    product_object = product_detail_template['product']
    product_object.pop('product_id', None)  # 移除product_id以创建新商品
    
    if 'comp_key_value_map' not in product_object:
        return None
    
    comp_map = product_object['comp_key_value_map']
    
    # 更新商品名称、价格和时间
    current_timestamp = int(time.time())
    comp_map['productName'] = new_data["团购标题"]
    
    # 更新售价和原价
    actual_amount = int(new_data["售价"] * 100)  # 转换为分
    origin_amount = int(new_data["原价"] * 100)  # 转换为分
    comp_map['actualAmount'] = str(actual_amount)
    comp_map['originAmount'] = str(origin_amount)
    
    sold_start_time = str(current_timestamp)
    sold_end_time = str(current_timestamp + 90 * 24 * 3600)
    comp_map['auto_renew-sold_end_time-sold_start_time'] = json.dumps({
        "soldStartTime": sold_start_time,
        "soldEndTime": sold_end_time,
        "autoRenew": True,
        "soldTimeType": 1
    })
    
    log_func(f"✅ 商品名称已更新为: {comp_map['productName']}")
    log_func(f"✅ 售价已更新为: {new_data['售价']}元 (原价: {new_data['原价']}元)")
    
    # 强制设置为"不需要"顾客信息
    comp_map['customer_reserved_info-real_name_info'] = '{"customerReservedInfo":{"allow":false},"realNameInfo":{"enable":false}}'
    log_func("✅ 已强制将 '顾客信息设置' 修改为 '不需要'。")
    
    # 图片保持不变（复用模板商品的图片）
    log_func("✅ 图片链接保持原样（复用模板商品图片）。")
    
    # 更新 commodity 字段中的价格（这是网页端API真正读取的原价）
    # 重要：根据commodity_type构建正确的结构
    try:
        commodity_type = new_data.get("commodity_type", "网费")
        log_func(f"--- [Commodity更新] 目标类型: {commodity_type}")
        
        # 根据类型构建commodity结构
        if commodity_type == "网费":
            # 网费类型：简单结构，不需要服务时长等字段
            # 注意：price必须是字符串格式！
            
            # 确定适用人群
            member_type = new_data.get("member_type", "不限制")
            if member_type == "新客":
                suitable_group_key = 2
                suitable_group_value = "本店新会员"
            elif member_type == "老客":
                suitable_group_key = 3
                suitable_group_value = "本店老会员"
            else:
                suitable_group_key = 1
                suitable_group_value = "不限制"
            
            commodity_obj = [{
                "group_name": "网费",
                "total_count": 1,
                "option_count": 1,
                "item_list": [{
                    "count": "1",
                    "count-unit": json.dumps({"count": 1, "unit": "FEN"}, ensure_ascii=False),
                    "includeMeal": json.dumps({"value": False}, ensure_ascii=False),
                    "itemOpticalItemClassify": json.dumps({"value": 1, "label": "网费服务", "isCustom": None}, ensure_ascii=False),
                    "itemSuitableGroup": json.dumps({"key": suitable_group_key, "value": suitable_group_value}, ensure_ascii=False),
                    "name": "网费",
                    "price": str(origin_amount),  # 必须是字符串！
                    "unit": "FEN"
                }]
            }]
            log_func(f"✅ 已构建网费类型commodity结构，原价: {origin_amount/100}元，适用人群: {suitable_group_value}")
        else:
            # 包时类型：需要保留服务时长等字段
            commodity_str = comp_map.get('commodity')
            if commodity_str:
                log_func(f"--- [Commodity更新] 原始commodity长度: {len(commodity_str)} 字符")
                commodity_obj = json.loads(commodity_str)
                log_func(f"--- [Commodity更新] 解析后有 {len(commodity_obj)} 个group")
                
                if commodity_obj and len(commodity_obj) > 0:
                    # 只保留第一个group
                    first_group = commodity_obj[0]
                    log_func(f"--- [Commodity更新] 第一个group有 {len(first_group.get('item_list', []))} 个item")
                    
                    if 'item_list' in first_group and len(first_group['item_list']) > 0:
                        # 只保留第一个item
                        first_item = first_group['item_list'][0]
                        
                        # 更新原价
                        old_price = first_item.get('price')
                        first_item['price'] = origin_amount
                        
                        # 更新名称
                        first_group['group_name'] = commodity_type
                        first_item['name'] = commodity_type
                        
                        # 只保留第一个item
                        first_group['item_list'] = [first_item]
                        first_group['total_count'] = 1
                        first_group['option_count'] = 1
                        
                        log_func(f"✅ 已简化并更新 commodity: price {old_price} → {origin_amount} ({origin_amount/100}元)")
                    
                    # 只保留第一个group
                    commodity_obj = [first_group]
            else:
                log_func("[Warning] 模板中没有找到 commodity 字段！")
                commodity_obj = None
        
        if commodity_obj:
            comp_map['commodity'] = json.dumps(commodity_obj, ensure_ascii=False)
            log_func(f"--- [Commodity更新] 最终commodity长度: {len(comp_map['commodity'])} 字符")
    except Exception as e:
        log_func(f"[Warning] 更新 commodity 价格时出错: {e}")
        import traceback
        log_func(f"详细错误: {traceback.format_exc()}")
    
    # 强制使用固定的 poi_set_id
    fixed_poi_set_id = "7585041807923316776"
    product_object['extra_map'] = {
        "poi_set_id": fixed_poi_set_id,
        "poi_check_result": "",
        "boost_strategy": '{"ai_recommend_title":"","ai_recommend_title_source":""}'
    }
    log_func(f"✅ 已强制将 'extra_map' 设置为固定值，poi_set_id 为: {fixed_poi_set_id}")
    
    # 更新 SKU 价格（如果存在）
    if 'sku' in product_detail_template:
        sku_object = product_detail_template['sku']
        log_func(f"--- [SKU更新前] actual_amount: {sku_object.get('actual_amount')}, origin_amount: {sku_object.get('origin_amount')}")
        sku_object['actual_amount'] = actual_amount
        sku_object['origin_amount'] = origin_amount
        sku_object['sku_name'] = new_data["团购标题"]
        log_func(f"--- [SKU更新后] actual_amount: {sku_object.get('actual_amount')}, origin_amount: {sku_object.get('origin_amount')}")
        log_func(f"✅ SKU 价格已同步更新: 售价={actual_amount/100}元, 原价={origin_amount/100}元")
    else:
        log_func("[Warning] product_detail_template 中没有 'sku' 字段")
    
    # 构建最终payload
    final_payload = {
        "product_detail": product_detail_template,
        "save_product_draft_cache_type": 4,
        "product_cache_scene": 1,
        "version_info": {
            "Enable": True,
            "VersionName": "1.0.8"
        }
    }
    
    # 打印关键价格信息用于调试
    log_func("--- [价格信息检查] ---")
    log_func(f"Product actualAmount: {comp_map.get('actualAmount')}")
    log_func(f"Product originAmount: {comp_map.get('originAmount')}")
    
    # 打印commodity字段中的价格
    try:
        commodity_str = comp_map.get('commodity')
        if commodity_str:
            commodity_obj = json.loads(commodity_str)
            log_func(f"Commodity 结构: {len(commodity_obj)} 个group")
            for idx, group in enumerate(commodity_obj):
                if 'item_list' in group:
                    for item_idx, item in enumerate(group['item_list']):
                        log_func(f"  Group[{idx}].item[{item_idx}].price = {item.get('price')}")
    except:
        pass
    
    if 'sku' in product_detail_template:
        log_func(f"SKU actual_amount: {product_detail_template['sku'].get('actual_amount')}")
        log_func(f"SKU origin_amount: {product_detail_template['sku'].get('origin_amount')}")
    
    return final_payload

def _create_product_web(session, product_payload, root_life_account_id, log_func):
    """通过网页端API创建商品"""
    log_func("--- [网页端] 正在发送创建商品请求... ---")
    url = "https://life.douyin.com/life/tobias/product/save/"
    params = {'root_life_account_id': root_life_account_id}
    
    # 打印完整的请求payload用于调试
    log_func("--- [完整请求Payload] ---")
    payload_str = json.dumps(product_payload, ensure_ascii=False, indent=2)
    # 打印完整payload，不截断（用于调试原价问题）
    log_func(payload_str)
    log_func("-" * 60)
    
    try:
        response = session.post(url, params=params, data=json.dumps(product_payload), timeout=20)
        response.raise_for_status()
        result = response.json()
        log_func(f"服务器响应: {json.dumps(result, ensure_ascii=False, indent=2)}")
        
        if result.get('status_code') == 0:
            product_id = result.get('product_id') or result.get('product', {}).get('product_id')
            if product_id and product_id != "0":
                log_func(f"[SUCCESS] 商品创建成功！Product ID: {product_id}")
                return product_id, None
        
        return None, result.get('status_msg', '未知API错误')
    except requests.exceptions.RequestException as e:
        return None, f"创建商品请求失败: {e}"

def _wait_for_product_approval(access_token, product_id, log_func, max_wait_time=60, check_interval=5):
    """等待商品审核通过"""
    log_func(f"--- 等待商品审核通过（最多等待{max_wait_time}秒）... ---")
    
    start_time = time.time()
    attempt = 0
    
    while time.time() - start_time < max_wait_time:
        attempt += 1
        log_func(f"第{attempt}次检查审核状态...")
        
        # 尝试获取商品详情
        product_details = get_douyin_product_details(access_token, product_id, log_func)
        
        if product_details:
            log_func(f"✅ 商品审核已通过！可以进行后续操作。")
            return True, product_details
        
        # 等待后再次检查
        if time.time() - start_time < max_wait_time:
            log_func(f"商品仍在审核中，{check_interval}秒后再次检查...")
            time.sleep(check_interval)
    
    log_func(f"[Warning] 等待超时（{max_wait_time}秒），商品可能仍在审核中。")
    return False, None

def create_product_via_web(cookie, csrf_token, root_life_account_id, template_product_id, new_data, target_poi_id, access_token, log_func):
    """使用网页端API创建商品（重创模式专用）- 复用模板图片，创建后自动修改POI ID"""
    log_func("========== 开始 重创 商品（网页端模式）==========")
    
    if not cookie or not csrf_token:
        log_func("[Error] 网页端Cookie或CSRF Token未配置，无法使用重创模式。")
        return None, "缺少网页端认证信息"
    
    if not template_product_id:
        log_func("[Error] 重创模式需要一个模板商品ID。")
        return None, "缺少模板商品ID"
    
    session = requests.Session()
    session.headers.update({
        'Accept': 'application/json, text/plain, */*',
        'Cookie': cookie,
        'Origin': 'https://life.douyin.com',
        'Referer': 'https://life.douyin.com/p/product/create',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36 Edg/141.0.0.0',
        'x-secsdk-csrf-token': csrf_token,
        'Content-Type': 'application/json;charset=UTF-8'
    })

    # 步骤1: 获取模板商品
    log_func(f"--- 步骤1: 获取模板商品 (ID: {template_product_id}) ---")
    product_detail_template, error = _get_product_template_web(session, template_product_id, root_life_account_id, log_func)
    if error:
        log_func(f"[FAILURE] 获取模板失败: {error}")
        return None, error

    # 步骤2: 构建payload（复用模板图片，使用固定POI ID）
    log_func("--- 步骤2: 构建商品payload（复用图片，固定POI ID） ---")
    product_payload = _build_web_product_payload_from_template(product_detail_template, new_data, log_func)
    if not product_payload:
        log_func("[FAILURE] 构建payload失败")
        return None, "构建payload失败"
    
    # 步骤3: 创建商品
    log_func("--- 步骤3: 创建新商品 ---")
    new_product_id, error = _create_product_web(session, product_payload, root_life_account_id, log_func)
    if error:
        log_func(f"[FAILURE] 商品 '{new_data['团购标题']}' 创建失败: {error}")
        return None, error
    
    log_func(f"✅ 商品创建成功！新商品ID: {new_product_id}")
    
    # 步骤4: 等待商品审核通过
    log_func(f"--- 步骤4: 等待商品审核通过 ---")
    approval_success, full_product_data = _wait_for_product_approval(access_token, new_product_id, log_func, max_wait_time=60, check_interval=5)
    
    if not approval_success or not full_product_data:
        log_func("[Warning] 商品可能仍在审核中，无法立即修改POI ID。")
        log_func("[Info] 商品已创建成功，但POI ID为固定值。请稍后手动修改或等待审核通过后重新运行。")
        return new_product_id, None
    
    # 步骤5: 使用开放平台API修改POI ID到目标门店
    log_func(f"--- 步骤5: 修改POI ID到目标门店 (POI ID: {target_poi_id}) ---")
    
    try:
        product_to_save = full_product_data.get('product')
        sku_to_save = full_product_data.get('skus', [{}])[0] if full_product_data.get('skus') else full_product_data.get('sku')
        
        if not product_to_save or not sku_to_save:
            log_func("[Warning] 商品数据不完整，POI ID可能未更新。")
            return new_product_id, None
        
        # 更新POI ID到目标门店
        product_to_save['pois'] = [{"poi_id": str(target_poi_id)}]
        extra_obj = json.loads(product_to_save.get("extra", "{}"))
        extra_obj['poi_set_id'] = str(target_poi_id)
        product_to_save['extra'] = json.dumps(extra_obj)
        
        # 确保所有必填字段存在
        log_func("正在检查并补充必填字段...")
        
        # 1. product 必填字段
        if "attr_key_value_map" not in product_to_save:
            product_to_save["attr_key_value_map"] = {}
        
        # RefundPolicy（退款政策）
        if "RefundPolicy" not in product_to_save["attr_key_value_map"]:
            product_to_save["attr_key_value_map"]["RefundPolicy"] = "2"
            log_func("已添加缺失的 RefundPolicy 字段")
        
        # Notification（使用须知）
        if "Notification" not in product_to_save["attr_key_value_map"]:
            notification_content = [
                {"title": "使用须知", "content": "请按照商家规定使用"},
                {"title": "限购说明", "content": "每人限购1份"},
                {"title": "有效期", "content": "购买后30日内有效"}
            ]
            product_to_save["attr_key_value_map"]["Notification"] = json.dumps(notification_content, ensure_ascii=False)
            log_func("已添加缺失的 Notification 字段")
        
        # Description（商品描述）
        if "Description" not in product_to_save["attr_key_value_map"]:
            product_to_save["attr_key_value_map"]["Description"] = json.dumps(["适用区域: 全场通用"], ensure_ascii=False)
            log_func("已添加缺失的 Description 字段")
        
        # 2. sku 必填字段
        if "attr_key_value_map" not in sku_to_save:
            sku_to_save["attr_key_value_map"] = {}
        
        # use_type（使用类型）
        if "use_type" not in sku_to_save.get("attr_key_value_map", {}):
            sku_to_save["attr_key_value_map"]["use_type"] = "1"
            log_func("已添加缺失的 use_type 字段")
        
        log_func(f"正在将POI ID从固定值更新为目标门店: {target_poi_id}")
        
        # 构建保存请求
        save_payload = {
            "account_id": str(DOUYIN_ACCOUNT_ID),
            "product": product_to_save,
            "sku": sku_to_save,
            "poi_ids": [str(target_poi_id)],
            "supplier_ext_ids": [str(target_poi_id)]
        }
        
        headers = {"Content-Type": "application/json", "access-token": access_token}
        response = requests.post(DOUYIN_PRODUCT_SAVE_URL, headers=headers, json=save_payload, timeout=20)
        response.raise_for_status()
        response_data = response.json()
        
        if response_data.get('data', {}).get('error_code') == 0:
            log_func(f"✅ POI ID已成功更新到目标门店！")
        else:
            log_func(f"[Warning] POI ID更新失败: {response_data.get('data', {}).get('description', '未知错误')}")
    
    except Exception as e:
        log_func(f"[Warning] 更新POI ID时出错: {e}")
    
    log_func(f"[SUCCESS] 商品 '{new_data['团购标题']}' 重创完成！Product ID: {new_product_id}")
    return new_product_id, None

def update_douyin_product(access_token, template_product_id, new_data, log_func, mode="修改", image_dir=None, target_poi_id=None):
    log_func(f"========== 开始 {mode} 商品 ==========")

    if mode == "重创":
        if not target_poi_id:
            log_func("[Error] 重创模式下需要一个目标门店ID(target_poi_id)，但未提供。")
            return False, "重创模式缺少目标门店ID"
        try:
            with open('商品完整内容.json', 'r', encoding='utf-8') as f:
                template_data = json.load(f)
            full_product_data = template_data['data']['product_onlines'][0]
            log_func("成功从 '商品完整内容.json' 加载模板。")
        except Exception as e:
            log_func(f"[Error] 加载模板文件 '商品完整内容.json' 失败: {e}")
            return False, "加载模板文件失败"
    else: # 修改模式
        if not template_product_id:
             log_func("[Error] 修改模式下需要一个模板商品ID，但未提供。")
             return False, "修改模式缺少模板ID"
        full_product_data = get_douyin_product_details(access_token, template_product_id, log_func)

    if not full_product_data: return False, "获取模板商品详情失败"

    try:
        product_to_save = full_product_data.get('product')
        sku_to_save = full_product_data.get('skus')[0] if full_product_data.get('skus') else full_product_data.get('sku')
        if not product_to_save or not sku_to_save: return False, f"模板商品数据不完整"

        log_func(f"使用模板进行{mode}操作...")
        
        poi_ids_for_saving = []

        if mode == "重创":
            # --- 动态更新POI ID ---
            product_to_save['pois'] = [{"poi_id": str(target_poi_id)}]
            extra_obj = json.loads(product_to_save.get("extra", "{}"))
            extra_obj['poi_set_id'] = str(target_poi_id)
            product_to_save['extra'] = json.dumps(extra_obj)
            poi_ids_for_saving.append(str(target_poi_id))
            log_func(f"商品POI ID已更新为: {target_poi_id}")

            # --- 动态更新用户类型 (commodity) ---
            member_type = new_data.get("member_type")
            if member_type:
                try:
                    commodity_str = sku_to_save['attr_key_value_map']['commodity']
                    commodity_obj = json.loads(commodity_str)
                    
                    # 更新适用人群
                    member_type_map = {
                        "新客": '{"key":2,"value":"仅限新客"}',
                        "老客": '{"key":3,"value":"仅限老客"}',
                        "不限制": '{"key":1,"value":"不限制"}'
                    }
                    if member_type in member_type_map and commodity_obj and commodity_obj[0].get('item_list'):
                        for item in commodity_obj[0]['item_list']:
                            for attr in item.get('attr_list', []):
                                if attr.get('attr_key') == 'item_suitable_group':
                                    attr['attr_value'] = member_type_map[member_type]
                                    log_func(f"商品适用人群已更新为: {member_type}")
                                    break
                    
                    # 更新套餐类型和原价
                    commodity_type = new_data.get("commodity_type")
                    origin_price = new_data.get("原价")
                    applicable_location = new_data.get("applicable_location")

                    if commodity_obj and commodity_obj[0].get('item_list'):
                        commodity_group = commodity_obj[0]
                        item_list_inner = commodity_group['item_list'][0]
                        
                        # 更新套餐类型和原价
                        if commodity_type and origin_price is not None:
                            log_func(f"准备根据类型 '{commodity_type}' 和原价 '{origin_price}' 更新commodity字段...")
                            commodity_group['group_name'] = commodity_type
                            item_list_inner['name'] = commodity_type
                            item_list_inner['price'] = int(origin_price * 100)
                            log_func(f"commodity内部的group_name和name已更新为'{commodity_type}'，price已更新为'{item_list_inner['price']}'")
                        
                        # 更新适用位置
                        if applicable_location:
                            for attr in item_list_inner.get('attr_list', []):
                                if attr.get('attr_key') == 'applicable_location':
                                    location_value = json.loads(attr['attr_value'])
                                    location_value['value'] = applicable_location
                                    attr['attr_value'] = json.dumps(location_value, ensure_ascii=False)
                                    log_func(f"商品适用位置已更新为: {applicable_location}")
                                    break
                        
                        # 更新项目分类
                        optical_classify_map = {
                            "包时": '{"key":2,"value":"上网包时类服务"}',
                            "网费": '{"key":1,"value":"网费服务"}'
                        }
                        if commodity_type in optical_classify_map:
                             for attr in item_list_inner.get('attr_list', []):
                                if attr.get('attr_key') == 'item_optical_item_classify':
                                    attr['attr_value'] = optical_classify_map[commodity_type]
                                    log_func(f"项目分类已根据套餐类型 '{commodity_type}' 更新。")
                                    break

                    sku_to_save['attr_key_value_map']['commodity'] = json.dumps(commodity_obj, ensure_ascii=False)
                except Exception as e:
                    log_func(f"[Warning] 更新商品 'commodity' 字段失败: {e}")
        else: # 修改模式
             extra_obj = json.loads(product_to_save.get("extra", "{}"))
             poi_set_id = extra_obj.get("poi_set_id")
             if not poi_set_id: return False, f"在 extra 字段中未找到 poi_set_id"
             poi_ids_for_saving.append(str(poi_set_id))

        if mode == "重创" and image_dir and new_data.get("matched_image"):
            image_filename = new_data["matched_image"]
            found_image_path = os.path.join(image_dir, image_filename)
            log_func(f"--- [新增套餐] 准备使用头图: {image_filename} ---")

            if os.path.exists(found_image_path):
                try:
                    original_image = Image.open(found_image_path)
                    poi_id_for_filename = str(json.loads(product_to_save.get("extra", "{}")).get("poi_set_id", "unknown_poi"))

                    log_func("--- 开始处理和上传图片 ---")
                    img_1_1 = center_crop_image(original_image, 1/1)
                    img_4_3 = center_crop_image(original_image, 4/3)
                    
                    url_1_1 = upload_to_r2(img_1_1, poi_id_for_filename, "1:1", log_func)
                    url_4_3 = upload_to_r2(img_4_3, poi_id_for_filename, "4:3", log_func)

                    if url_1_1 and url_4_3:
                        log_func("图片上传成功，正在更新商品头图信息...")
                        image_list_4_3 = [{"url": url_4_3}, {"url": url_4_3}]
                        product_to_save['attr_key_value_map']['image_list'] = json.dumps(image_list_4_3, ensure_ascii=False)
                        image_list_1v1 = [{"url": url_1_1}, {"url": url_4_3}]
                        product_to_save['attr_key_value_map']['image_1v1_list'] = json.dumps(image_list_1v1, ensure_ascii=False)
                        log_func("商品头图信息已更新。")
                    else:
                        log_func("[Warning] 图片上传失败，将不带头图创建套餐。")
                except Exception as img_e:
                    log_func(f"[Error] 处理或上传图片 '{image_filename}' 时出错: {img_e}，将不带头图创建套餐。")
            else:
                log_func(f"[Error] 在目录 '{image_dir}' 中未找到指定的图片文件 '{image_filename}'，将不带头图创建套餐。")

        product_to_save["product_name"] = new_data["团购标题"]
        notification_content = [{"title": "使用须知", "content": new_data['团单备注']}, {"title": "限购说明", "content": new_data['限购']}, {"title": "有效期", "content": f"购买后{new_data['有效期']}内有效"}]
        product_to_save['attr_key_value_map']['Notification'] = json.dumps(notification_content, ensure_ascii=False)
        product_to_save['attr_key_value_map']['Description'] = json.dumps([f"适用区域: {new_data['可用区域']}"], ensure_ascii=False)
        if "RefundPolicy" not in product_to_save["attr_key_value_map"]: product_to_save["attr_key_value_map"]["RefundPolicy"] = "2"

        sku_to_save["sku_name"] = new_data["团购标题"]
        sku_to_save["actual_amount"] = int(new_data["售价"] * 100)
        # 同时更新原价
        if new_data.get("原价"):
            sku_to_save["origin_amount"] = int(new_data["原价"] * 100)
        elif "origin_amount" not in sku_to_save:
             sku_to_save["origin_amount"] = int(new_data["售价"] * 100) # 如果没有提供原价，默认等于售价

        if "use_type" not in sku_to_save.get("attr_key_value_map", {}):
            if "attr_key_value_map" not in sku_to_save: sku_to_save["attr_key_value_map"] = {}
            sku_to_save["attr_key_value_map"]["use_type"] = "1"
        
        if mode == "重创":
            if "product_id" in product_to_save: del product_to_save["product_id"]
            if "sku_id" in sku_to_save: del sku_to_save["sku_id"]
            log_func("操作模式为“重创”，已移除 product_id 和 sku_id。")

        save_payload = {"account_id": str(DOUYIN_ACCOUNT_ID), "product": product_to_save, "sku": sku_to_save, "poi_ids": poi_ids_for_saving, "supplier_ext_ids": poi_ids_for_saving}
        log_func(f"准备发送 {mode} 请求...")
        log_func(f"--- [API Request Payload] ---\n{json.dumps(save_payload, ensure_ascii=False, indent=2)}\n" + "-"*30)
        
        headers = {"Content-Type": "application/json", "access-token": access_token}
        response = requests.post(DOUYIN_PRODUCT_SAVE_URL, headers=headers, json=save_payload, timeout=20)
        response.raise_for_status()
        response_data = response.json()

        if response_data.get('data', {}).get('error_code') == 0:
            log_func(f"[SUCCESS] 商品 '{new_data['团购标题']}' {mode}成功!")
            return True, ""
        else:
            reason = response_data.get('data', {}).get('description', 'API返回未知错误')
            log_func(f"[FAILURE] 商品 '{new_data['团购标题']}' {mode}失败: {reason}")
            return False, reason
    except Exception as e:
        log_func(f"处理商品时发生意外错误: {e}\n{traceback.format_exc()}")
        return False, f"意外错误: {e}"

def extract_cells_with_formatting(file_path, log_func):
    """提取Excel单元格内容和格式化信息（颜色等）"""
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        
        log_func("正在提取Excel单元格格式化信息...")
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        cells_data = []
        for row_idx in range(1, ws.max_row + 1):
            row_data = {}
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 跳过合并单元格
                if isinstance(cell, MergedCell):
                    continue
                    
                value = cell.value
                
                # 获取背景颜色
                fill = cell.fill
                bg_color = None
                try:
                    if hasattr(fill, 'start_color') and fill.start_color.rgb and str(fill.start_color.rgb) != '00000000':
                        bg_color = f"#{fill.start_color.rgb}"
                except:
                    pass
                
                # 将列索引转换为列字母
                col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                col_index = col_idx
                
                row_data[f"col_{col_index}"] = {
                    "value": value,
                    "bg_color": bg_color,
                    "position": f"R{row_idx}C{col_index}",
                    "col_letter": col_letter
                }
            
            if any(cell_info["value"] for cell_info in row_data.values() if cell_info["value"] is not None):
                cells_data.append({
                    "row": row_idx,
                    "cells": row_data
                })
        
        log_func(f"成功提取 {len(cells_data)} 行数据")
        return cells_data
    
    except Exception as e:
        log_func(f"[Error] 提取Excel格式化信息失败: {e}")
        import traceback
        log_func(f"详细错误信息: {traceback.format_exc()}")
        return None

def build_llm_prompt_for_table_parsing(cells_data, log_func):
    """构建用于智能表格解析的LLM Prompt"""
    
    # 将表格数据转换为文本表示，包含颜色信息
    table_text = "表格数据（包含位置、行号、列号、值和背景颜色）:\n\n"
    
    for row_info in cells_data:
        row_num = row_info["row"]
        table_text += f"第{row_num}行: "
        row_parts = []
        
        for col_key in sorted(row_info["cells"].keys()):
            cell_info = row_info["cells"][col_key]
            value = cell_info["value"]
            bg_color = cell_info["bg_color"]
            position = cell_info["position"]
            
            if value is not None and str(value).strip():
                if bg_color:
                    row_parts.append(f"[{position}]'{value}'({bg_color})")
                else:
                    row_parts.append(f"[{position}]'{value}'")
        
        if row_parts:
            table_text += " | ".join(row_parts) + "\n"
    
    prompt = f"""
# 任务：智能解析Excel表格数据

你是一个专业的Excel表格解析专家。请根据提供的表格数据，智能识别表头和数据，并理解业务意图。

## 表格数据：
{table_text}

## 解析要求：

1. **智能表头识别**：
   - 分析表格结构，确定哪一行是表头行
   - 表头通常包含"标题"、"价格"、"区域"、"限购"、"有效期"、"备注"等关键词
   - 如果表头被标记了特殊颜色（如黄色#FFFF00），优先考虑该行为表头

2. **数据行解析**：
   - 跳过表头，解析实际数据行
   - 将每行数据映射到标准字段：团购标题、售价、可用区域、限购、有效期、团单备注

3. **颜色意图识别**：
   - 重点关注背景颜色为黄色的单元格，这通常表示**修改意图**（intent: "modify"）
   - 如果整行都有黄色背景，则该行数据为修改操作
   - 如果没有特殊颜色标记，默认为新增操作（intent: "add"）

4. **数据标准化**：
   - 确保价格字段为数字格式
   - 清理和格式化文本内容
   - 填充缺失的字段（默认为空字符串）

## 返回格式：
请严格按照以下JSON格式返回结果，不要包含任何额外的解释：

```json
{{
  "header_row": 表头行号,
  "data": [
    {{
      "团购标题": "商品标题",
      "售价": 0.00,
      "可用区域": "区域描述",
      "限购": "限购说明",
      "有效期": "有效期",
      "团单备注": "备注内容",
      "intent": "modify"
    }}
  ],
  "analysis": {{
    "total_rows": 总行数,
    "header_columns": 表头列映射,
    "color_analysis": "颜色分析结果"
  }}
}}
```

**注意事项**：
- 如果无法识别表头，header_row设为null
- intent字段：如果单元格/行有特殊颜色标记则为"modify"，否则为"add"
- 价格必须是数字格式，无法转换时设为0.0
"""
    
    return prompt

def intelligent_load_excel_data(file_path, log_func, cache):
    """使用LLM智能解析Excel数据（简化版）"""
    try:
        log_func("正在使用LLM智能解析Excel表格...")
        
        # 1. 首先用pandas读取Excel，智能跳过空行
        df_list = []
        best_df_info = None
        
        # 先读取完整的Excel文件来分析结构
        try:
            df_raw = pd.read_excel(file_path, engine='openpyxl', header=None)
            log_func(f"原始Excel文件形状: {df_raw.shape}")
            
            # 智能寻找真正的数据开始位置
            data_start_row = None
            for i in range(min(10, len(df_raw))):  # 检查前10行
                row_data = df_raw.iloc[i]
                # 计算非空单元格数量
                non_empty_count = sum(1 for val in row_data if pd.notna(val) and str(val).strip())
                
                # 如果这一行有足够的数据，认为是数据开始行
                if non_empty_count >= 3:  # 至少3列有数据
                    data_start_row = i
                    log_func(f"检测到数据开始行: {i} (非空单元格: {non_empty_count})")
                    break
            
            if data_start_row is None:
                log_func("未找到有效数据行，使用默认header=1")
                data_start_row = 1
            
            # 从找到的数据开始行读取几次，确定最佳表头位置
            for header_offset in [0, 1, 2]:  # 在数据开始位置基础上偏移
                actual_header_row = data_start_row + header_offset
                if actual_header_row >= len(df_raw):
                    continue
                    
                try:
                    df = pd.read_excel(file_path, engine='openpyxl', header=actual_header_row)
                    if not df.empty and len(df.columns) > 3:  # 确保有足够的列
                        # 检查是否有有效的列名（非Unnamed且有意义）
                        valid_columns = [col for col in df.columns if not str(col).startswith('Unnamed')]
                        
                        df_info = {
                            'header_row': actual_header_row,
                            'data_start_row': data_start_row,
                            'dataframe': df,
                            'columns': list(df.columns),
                            'valid_column_count': len(valid_columns),
                            'valid_columns': valid_columns
                        }
                        df_list.append(df_info)
                        
                        log_func(f"读取Excel（header={actual_header_row}，数据开始={data_start_row}），共{len(df)}行，{len(df.columns)}列，有效列名{len(valid_columns)}个")
                        
                        # 如果找到有意义的列名，优先选择
                        if valid_columns and len(valid_columns) >= 3:
                            best_df_info = df_info
                            log_func(f"找到有效列名，跳过其他选项")
                            break
                            
                except Exception as e:
                    log_func(f"读取Excel（header={actual_header_row}）失败: {e}")
            
            # 如果没有找到有意义的列名，使用数据行数最多的
            if not best_df_info and df_list:
                best_df_info = max(df_list, key=lambda x: len(x['dataframe']))
                log_func(f"未找到有效列名，选择数据最多的选项（header={best_df_info['header_row']}）")
            
        except Exception as e:
            log_func(f"分析Excel结构失败: {e}，使用传统方法")
            # 回退到原始方法
            for header_row in [0, 1, 2, 3]:
                try:
                    df = pd.read_excel(file_path, engine='openpyxl', header=header_row)
                    if not df.empty and len(df.columns) > 3:
                        df_list.append({
                            'header_row': header_row,
                            'dataframe': df,
                            'columns': list(df.columns)
                        })
                except:
                    continue
        
        if not df_list:
            log_func("无法读取Excel文件，回退到传统方式")
            return load_excel_data_fallback(file_path, log_func)
        
        # 2. 如果已经有best_df_info，直接使用；否则从列表中选择最佳的
        if not best_df_info:
            if len(df_list) > 1:
                log_func("检测到多个可能的表头位置，使用LLM智能选择...")
                
                # 构建多个选项的描述
                options_text = ""
                for i, df_info in enumerate(df_list):
                    valid_cols = len([col for col in df_info['columns'] if not str(col).startswith('Unnamed')])
                    options_text += f"选项{i+1}（header={df_info['header_row']}，有效列名={valid_cols}个）: {df_info['columns']}\n"
                
                prompt = f"""
# 任务：智能识别Excel表格的最佳数据结构

Excel文件中有多个可能的表头位置，请选择最合理的一个。

## 可能的表头选项：
{options_text}

## 任务要求：
1. 优先选择有有效列名的选项（非Unnamed列）
2. 如果都有Unnamed列，选择数据行数最多的选项
3. 寻找包含以下关键词的列：标题、商品、名称、价格、区域、限购、有效期、备注

## 返回格式：
只返回一个数字（1-{len(df_list)}），表示选择的选项编号。

例如：如果选项2最合理，返回"2"
"""
                
                if llm_client:
                    try:
                        response = llm_client.chat.completions.create(
                            model=LLM_MODEL_ID,
                            messages=[{'role': 'user', 'content': prompt}],
                            stream=False
                        )
                        
                        choice_text = response.choices[0].message.content.strip()
                        log_func(f"--- [LLM Raw Response for Table Structure Selection] ---\n{choice_text}\n" + "-"*30)
                        
                        # 尝试提取数字
                        numbers = re.findall(r'\d+', choice_text)
                        if numbers:
                            choice_num = int(numbers[0]) - 1  # 转换为0-based索引
                            
                            if 0 <= choice_num < len(df_list):
                                best_df_info = df_list[choice_num]
                                log_func(f"LLM选择选项{choice_num+1}（header={best_df_info['header_row']}）")
                            else:
                                log_func(f"LLM返回数字超出范围 {numbers[0]}，使用默认选项")
                        else:
                            log_func(f"LLM返回中未找到数字 '{choice_text}'，分析内容...")
                            # 如果所有表头都是Unnamed，选择数据行数最多的
                            best_df_info = max(df_list, key=lambda x: len(x['dataframe']))
                            log_func(f"基于数据行数选择选项{df_list.index(best_df_info)+1}（header={best_df_info['header_row']}）")
                            
                    except Exception as e:
                        log_func(f"LLM选择失败: {e}，使用默认选项")
                else:
                    log_func("LLM客户端不可用，使用默认选项")
            else:
                best_df_info = df_list[0]  # 只有一个选项时直接使用
        
        # 3. 使用选定的DataFrame进行字段映射
        selected_df = best_df_info['dataframe']
        columns = best_df_info['columns']
        
        log_func(f"选定数据结构：header={best_df_info['header_row']}, 列数={len(columns)}")
        
        # 4. 构建字段映射Prompt
        mapping_prompt = f"""
# 任务：智能映射Excel列到标准字段

Excel列名: {columns}

需要映射到的标准字段：
- 团购标题（商品名称）
- 售价（数字价格）
- 可用区域（适用区域）
- 限购（购买限制）
- 有效期（使用期限）
- 团单备注（说明备注）

## 任务要求：
1. 分析每个Excel列的内容，判断它对应哪个标准字段
2. 如果某个标准字段在Excel中找不到对应列，设为null
3. 返回严格的JSON映射对象

## 返回格式：
```json
{{
  "团购标题": "对应的Excel列名或null",
  "售价": "对应的Excel列名或null",
  "可用区域": "对应的Excel列名或null",
  "限购": "对应的Excel列名或null",
  "有效期": "对应的Excel列名或null",
  "团单备注": "对应的Excel列名或null"
}}
```
"""
        
        column_mapping = None
        
        # 检查是否所有列名都是"Unnamed"
        unnamed_count = sum(1 for col in columns if str(col).startswith('Unnamed'))
        if unnamed_count == len(columns):
            log_func("检测到所有列头都是Unnamed，使用数据内容分析...")
            
            # 获取前3行数据样本进行分析
            sample_data = []
            for i in range(min(3, len(selected_df))):
                row_data = [str(selected_df.iloc[i, j]) if j < len(selected_df.columns) else "" for j in range(len(columns))]
                sample_data.append(row_data)
            
            # 构建基于数据内容的分析Prompt
            content_mapping_prompt = f"""
# 任务：基于数据内容智能映射Excel列

Excel列: {columns}

前3行数据样本:
{chr(10).join([f'第{i+1}行: {row}' for i, row in enumerate(sample_data)])}

## 任务要求：
基于数据内容判断每列的含义，并映射到以下标准字段：
- 团购标题（商品名称，通常包含商品、套餐、标题等关键词）
- 售价（数字价格，可能是小数或整数）
- 可用区域（描述性文本，位置、区域相关）
- 限购（购买限制说明）
- 有效期（时间期限描述）
- 团单备注（补充说明）

## 返回格式：
只返回JSON对象，不要其他解释：
```json
{{
  "团购标题": "列索引(0-11)",
  "售价": "列索引或null",
  "可用区域": "列索引或null",
  "限购": "列索引或null",
  "有效期": "列索引或null",
  "团单备注": "列索引或null"
}}
```
"""
            
            if llm_client:
                try:
                    response = llm_client.chat.completions.create(
                        model=LLM_MODEL_ID,
                        messages=[{'role': 'user', 'content': content_mapping_prompt}],
                        stream=False
                    )
                    
                    response_text = response.choices[0].message.content.strip()
                    log_func(f"--- [LLM Raw Response for Content-Based Mapping] ---\n{response_text}\n" + "-"*30)
                    json_match = re.search(r'\{[\s\S]*\}', response_text)
                    if json_match:
                        cleaned_response = json_match.group(0)
                        column_mapping = json.loads(cleaned_response)
                    else:
                        raise json.JSONDecodeError("在LLM响应中未找到JSON对象", response_text, 0)
                    log_func(f"基于数据内容的LLM映射结果: {column_mapping}")
                    
                    # 转换列索引为实际列名
                    index_mapping = {}
                    for field, idx_str in column_mapping.items():
                        if idx_str and str(idx_str).isdigit():
                            idx = int(idx_str)
                            if 0 <= idx < len(columns):
                                index_mapping[field] = columns[idx]
                            else:
                                index_mapping[field] = None
                        else:
                            index_mapping[field] = None
                    
                    column_mapping = index_mapping
                    log_func(f"转换为列名映射: {column_mapping}")
                    
                except Exception as e:
                    log_func(f"基于数据内容的LLM映射失败: {e}")
        
        # 如果还是失败，尝试原始的列名映射
        if not column_mapping and llm_client:
            try:
                response = llm_client.chat.completions.create(
                    model=LLM_MODEL_ID,
                    messages=[{'role': 'user', 'content': mapping_prompt}],
                    stream=False
                )
                
                response_text = response.choices[0].message.content.strip()
                log_func(f"--- [LLM Raw Response for Column Name Mapping] ---\n{response_text}\n" + "-"*30)
                json_match = re.search(r'\{[\s\S]*\}', response_text)
                if json_match:
                    cleaned_response = json_match.group(0)
                    column_mapping = json.loads(cleaned_response)
                else:
                    raise json.JSONDecodeError("在LLM响应中未找到JSON对象", response_text, 0)
                log_func(f"LLM字段映射结果: {column_mapping}")
                
            except Exception as e:
                log_func(f"LLM字段映射失败: {e}")
        
        # 5. 如果LLM映射失败，使用启发式映射
        if not column_mapping:
            log_func("LLM完全失败，使用强制内容分析...")
            # 强制基于数据内容进行智能映射
            column_mapping = force_content_based_mapping(selected_df, columns, log_func)
            log_func(f"强制内容映射结果: {column_mapping}")
        
        # 6. 应用映射并清理数据
        mapped_data = apply_column_mapping(selected_df, column_mapping, log_func)
        
        log_func(f"智能解析完成，共{mapped_data['record_count']}条记录")
        return mapped_data['records']
        
    except Exception as e:
        log_func(f"[Error] 智能解析失败: {e}，回退到传统方式")
        return load_excel_data_fallback(file_path, log_func)

def force_content_based_mapping(df, columns, log_func):
    """强制基于数据内容进行字段映射"""
    log_func("执行强制内容映射...")
    
    # 获取前5行数据样本
    sample_data = []
    for i in range(min(5, len(df))):
        row_data = []
        for j in range(len(columns)):
            value = df.iloc[i, j] if j < len(df.columns) else ""
            row_data.append(str(value))
        sample_data.append(row_data)
    
    # 分析每列的数据特征
    column_analysis = {}
    for col_idx, col_name in enumerate(columns):
        if col_idx >= len(sample_data[0]):
            continue
            
        values = [row[col_idx] for row in sample_data if col_idx < len(row)]
        non_empty_values = [v for v in values if v and str(v).strip() and str(v) != 'nan']
        
        analysis = {
            'column_name': col_name,
            'total_values': len(values),
            'non_empty_count': len(non_empty_values),
            'sample_values': non_empty_values[:3],  # 前3个非空值
            'is_numeric': True,  # 默认假设是数字
            'is_text': True,     # 默认假设是文本
            'looks_like_title': False,
            'looks_like_price': False,
            'looks_like_area': False,
            'looks_like_limit': False,
            'looks_like_validity': False,
            'looks_like_remark': False
        }
        
        # 数字特征分析
        numeric_count = 0
        for v in non_empty_values:
            try:
                # 尝试转换为数字
                if isinstance(v, str):
                    # 清理价格格式
                    clean_v = v.replace('¥', '').replace(',', '').strip()
                    if clean_v and clean_v.replace('.', '').replace('-', '').isdigit():
                        numeric_count += 1
                elif isinstance(v, (int, float)) and not pd.isna(v):
                    numeric_count += 1
            except:
                pass
        
        analysis['is_numeric'] = numeric_count > len(non_empty_values) * 0.6  # 60%以上是数字
        
        # 文本特征分析
        text_samples = ' '.join(non_empty_values).lower()
        
        # 标题特征：包含商品相关词汇
        title_keywords = ['套餐', '商品', '网费', '包时', '会员', '新客', '老客', '特价', '优惠', '活动', '团购']
        analysis['looks_like_title'] = any(keyword in text_samples for keyword in title_keywords)
        
        # 价格特征：数字且包含价格相关词汇
        price_keywords = ['元', '价格', '售价', '优惠价']
        analysis['looks_like_price'] = analysis['is_numeric'] or any(keyword in text_samples for keyword in price_keywords)
        
        # 区域特征
        area_keywords = ['区域', '位置', '区域', '适用', '大厅', '包间', '单人', '双人']
        analysis['looks_like_area'] = any(keyword in text_samples for keyword in area_keywords)
        
        # 限购特征
        limit_keywords = ['限购', '限制', '购买', '次数', '人均']
        analysis['looks_like_limit'] = any(keyword in text_samples for keyword in limit_keywords)
        
        # 有效期特征
        validity_keywords = ['有效期', '期限', '天', '月', '日', '年']
        analysis['looks_like_validity'] = any(keyword in text_samples for keyword in validity_keywords)
        
        # 备注特征
        remark_keywords = ['备注', '说明', '备注', '注意', '须知', '使用']
        analysis['looks_like_remark'] = any(keyword in text_samples for keyword in remark_keywords)
        
        column_analysis[col_idx] = analysis
    
    # 根据特征进行字段映射
    mapping = {
        "团购标题": None,
        "售价": None,
        "可用区域": None,
        "限购": None,
        "有效期": None,
        "团单备注": None
    }
    
    # 映射逻辑
    for col_idx, analysis in column_analysis.items():
        col_name = analysis['column_name']
        
        # 标题映射 - 优先文本且有标题特征的
        if mapping["团购标题"] is None and (analysis['looks_like_title'] or (not analysis['is_numeric'] and len(analysis['sample_values']) > 0)):
            mapping["团购标题"] = col_name
            log_func(f"映射标题列: {col_name} (特征: 标题={analysis['looks_like_title']}, 数字={analysis['is_numeric']})")
            continue
            
        # 价格映射 - 优先数字特征
        if mapping["售价"] is None and analysis['looks_like_price']:
            mapping["售价"] = col_name
            log_func(f"映射价格列: {col_name} (特征: 价格={analysis['looks_like_price']}, 数字={analysis['is_numeric']})")
            continue
            
        # 区域映射 - 区域特征优先
        if mapping["可用区域"] is None and analysis['looks_like_area']:
            mapping["可用区域"] = col_name
            log_func(f"映射区域列: {col_name}")
            continue
            
        # 限购映射 - 限购特征优先
        if mapping["限购"] is None and analysis['looks_like_limit']:
            mapping["限购"] = col_name
            log_func(f"映射限购列: {col_name}")
            continue
            
        # 有效期映射 - 有效期特征优先
        if mapping["有效期"] is None and analysis['looks_like_validity']:
            mapping["有效期"] = col_name
            log_func(f"映射有效期列: {col_name}")
            continue
            
        # 备注映射 - 其他文本列
        if mapping["团单备注"] is None and not analysis['is_numeric'] and analysis['sample_values']:
            mapping["团单备注"] = col_name
            log_func(f"映射备注列: {col_name}")
            continue
    
    # 如果还有未映射的字段，尝试智能分配
    unmapped_fields = [field for field, col in mapping.items() if col is None]
    unmapped_cols = [col_idx for col_idx, analysis in column_analysis.items() if analysis['column_name'] not in mapping.values()]
    
    for field in unmapped_fields:
        if unmapped_cols:
            col_idx = unmapped_cols.pop(0)
            col_name = column_analysis[col_idx]['column_name']
            mapping[field] = col_name
            log_func(f"智能分配 {field} -> {col_name}")
    
    return mapping

def heuristic_column_mapping(columns):
    """启发式字段映射"""
    mapping = {
        "团购标题": None,
        "售价": None,
        "可用区域": None,
        "限购": None,
        "有效期": None,
        "团单备注": None
    }
    
    # 关键词映射
    keyword_map = {
        "团购标题": ["标题", "商品", "名称", "产品", "套餐", "title", "name"],
        "售价": ["售价", "价格", "金额", "price", "现价"],
        "可用区域": ["区域", "地区", "适用", "位置", "area", "location"],
        "限购": ["限购", "限制", "购买", "limit"],
        "有效期": ["有效期", "期限", "天数", "validity", "expire"],
        "团单备注": ["备注", "说明", "note", "remark", "描述", "description"]
    }
    
    for field, keywords in keyword_map.items():
        for col in columns:
            col_lower = str(col).lower()
            if any(keyword.lower() in col_lower for keyword in keywords):
                mapping[field] = col
                break
    
    return mapping

def apply_column_mapping(df, column_mapping, log_func):
    """应用列映射并清理数据"""
    records = []
    
    # 数据清理
    df_filled = df.fillna('')
    
    for _, row in df_filled.iterrows():
        record = {}
        for field, excel_col in column_mapping.items():
            if excel_col and excel_col in df.columns:
                value = row[excel_col]
                # 特殊处理价格字段
                if field == "售价" and value:
                    try:
                        # 尝试转换为数字
                        if isinstance(value, str):
                            # 移除货币符号和逗号
                            value = value.replace('¥', '').replace(',', '').strip()
                        record[field] = float(value)
                    except (ValueError, TypeError):
                        record[field] = 0.0
                else:
                    record[field] = str(value) if value else ""
            else:
                record[field] = "" if field != "售价" else 0.0
        
        # 只添加有实际内容的记录
        if any(str(v).strip() for v in record.values() if v != 0.0):
            records.append(record)
    
    return {
        "records": records,
        "record_count": len(records),
        "mapping_used": column_mapping
    }

def load_excel_data_fallback(file_path, log_func):
    """传统Excel加载方式作为回退方案"""
    try:
        df = pd.read_excel(file_path, engine='openpyxl', header=1)
        expected_columns = {df.columns[4]: '团购标题', df.columns[5]: '售价', df.columns[6]: '可用区域', df.columns[7]: '限购', df.columns[8]: '有效期', df.columns[9]: '团单备注'}
        df.rename(columns=expected_columns, inplace=True)
        df.fillna('', inplace=True)
        data = df.to_dict('records')
        log_func(f"[Success] 使用传统方式成功加载 {len(data)} 条Excel数据。")
        return data
    except Exception as e:
        log_func(f"[Error] 传统方式加载Excel文件也失败: {e}")
    return None

# 保持向后兼容的别名
def load_excel_data(file_path, log_func):
    """智能Excel数据加载函数（支持LLM增强）"""
    return intelligent_load_excel_data(file_path, log_func, {})

def match_products_with_llm(douyin_products, excel_data, log_func, cache):
    if not llm_client: return None
    log_func("--- 开始使用LLM智能匹配套餐 ---")
    # 准备更详细的数据给LLM，以提高匹配准确性
    douyin_product_details_for_llm = [
        {"name": p['name'], "price": p['price'], "origin_price": p.get('origin_price', '0.00')}
        for p in douyin_products
    ]
    excel_product_details_for_llm = [
        {"团购标题": p['团购标题'], "售价": p.get('售价', 0.0), "网费": p.get('网费', 0.0), "区域": p.get('可用区域', '')}
        for p in excel_data
    ]

    prompt = f"""
# 任务：智能匹配抖音商品与Excel商品

请为“抖音商品列表”中的每一个商品，在“Excel商品列表”中找到唯一且最精确的匹配项。

## 匹配原则 (极其重要):
1.  **核心匹配**: 首先必须根据商品的核心内容进行匹配。例如，抖音的“【新会员】108网费”应该匹配Excel中的“【新会员】108网费”。
2.  **价格验证**: 在核心内容匹配的基础上，必须严格比较价格。抖音商品的价格 (`price`) 必须与Excel中对应的“售价”**几乎完全相等**。
3.  **内容与价格结合**: 综合核心内容和价格进行双重验证。例如，一个抖音商品叫“【上午包】无烟区”，价格是10.5元，那么它应该匹配到Excel中“团购标题”为“【上午包】无烟区”且“售价”为10.5的那一行。
4.  **处理模糊情况**: 如果多个抖音商品（例如“【瞬影必杀券】100元网费（新会员）”和“【新鼠鼠券A】100元网费”）都能模糊匹配到同一个Excel项（例如“【新会员】108网费”），你需要根据**价格**来区分。如果价格也相似，则选择内容更接近的那个。如果无法区分，则可以将其中一个设为null。
5.  **找不到则为null**: 如果在Excel列表中找不到任何满足上述条件的匹配项，对应的值必须是 `null`。

## 数据列表:

### 抖音商品列表 (包含名称、现价、原价):
{json.dumps(douyin_product_details_for_llm, ensure_ascii=False, indent=2)}

### Excel商品列表 (包含团购标题、售价、网费、区域):
{json.dumps(excel_product_details_for_llm, ensure_ascii=False, indent=2)}

## 返回格式:
请严格按照以下JSON格式返回结果，键是抖音商品**完整的`name`**，值是匹配到的Excel商品**完整的`团购标题`**。不要包含任何额外的解释。

```json
{{
  "抖音商品名称1": "匹配到的Excel团购标题1",
  "抖音商品名称2": null,
  "抖音商品名称3": "匹配到的Excel团购标题3"
}}
```
"""
    
    cache_key = hashlib.sha256(prompt.encode('utf-8')).hexdigest()
    if cache_key in cache:
        log_func(f"[Cache Hit] 发现相同请求的缓存结果，直接使用缓存。")
        return cache[cache_key]

    log_func("[Cache Miss] 未找到缓存，发起新的LLM请求。")
    try:
        log_func("正在调用LLM API...")
        log_func(f"--- [DEBUG] 发送给LLM的Prompt ---\n{prompt[:500]}...\n" + "-"*30)
        response = llm_client.chat.completions.create(model=LLM_MODEL_ID, messages=[{'role': 'user', 'content': prompt}], stream=True)
        full_response = "".join(chunk.choices[0].delta.content for chunk in response if chunk.choices[0].delta.content)
        log_func(f"--- [LLM Raw Response for Product Matching] ---\n{full_response}\n" + "-"*30)
        json_match = re.search(r'\{[\s\S]*\}', full_response)
        if json_match:
            cleaned_response = json_match.group(0)
            match_result = json.loads(cleaned_response)
        else:
            raise json.JSONDecodeError("在LLM响应中未找到JSON对象", full_response, 0)
        log_func("[Success] LLM智能匹配成功！")
        cache[cache_key] = match_result
        log_func(f"已将本次结果存入缓存，Key: {cache_key[:10]}...")
        return match_result
    except Exception as e:
        log_func(f"[Error] LLM智能匹配过程中出错: {e}")
    return None

def analyze_text_for_actions(text_input, douyin_products, log_func, cache):
    if not llm_client:
        log_func("[Error] LLM客户端未初始化。")
        return None
        
    log_func("--- 开始使用LLM分析文本指令 ---")
    
    # 转换价格为数字，以便AI更好地处理
    simple_product_list_for_llm = []
    for p in douyin_products:
        try:
            price = float(p.get('price', 0))
            origin_price = float(p.get('origin_price', 0))
        except (ValueError, TypeError):
            price = 0.0
            origin_price = 0.0
        simple_product_list_for_llm.append({
            "name": p.get('name'),
            "price": price,
            "origin_price": origin_price
        })

    prompt = f"""
你是一个专业的抖音团购运营助理。请根据用户提供的文本指令和当前的抖音线上商品列表，分析出需要执行的新增、修改、下架操作。

# 当前线上商品列表 (包含名称、现价和原价):
{json.dumps(simple_product_list_for_llm, ensure_ascii=False, indent=2)}

# 用户指令:
---
{text_input}
---

# 任务要求:
1.  **分析指令**: 仔细阅读用户指令，识别出三种操作：`add` (新增), `update` (修改), `delete` (下架)。
    *   **特别注意：如果用户指令中明确包含“新建”或“新增”关键词，则应将所有内容都解析为 `add` 操作，不要尝试进行 `update` 或 `delete` 匹配。**

2.  **智能匹配商品 (核心任务)**:
    *   对于 `update` 和 `delete` 操作（**仅在指令不含“新建”或“新增”时执行**），你必须在“当前线上商品列表”中找到最相关的商品。匹配不应是简单的文本相等，而应是基于核心内容、现价 (`price`) 和原价 (`origin_price`) 的**智能模糊匹配**。
    *   **匹配示例**: 用户指令 `【9.9得60】换成【19.9得100】` 应该能准确匹配到线上商品 `{{"name": "【开业新会员】9.9得60网费", "price": 9.9, "origin_price": 60.0}}`，因为它的核心部分 `9.9得60` 与价格 `9.9` 和 `60.0` 高度相关。
    *   对于 `delete` 操作，如果指令只有价格（如 `59.9下架`），应根据 `price` 字段进行匹配。
    *   对于 `add` 操作，是全新的商品，不需要匹配。

3.  **提取并构建信息**:
    *   对于 `add` 操作，根据指令提取并构建一个完整商品信息对象。
        *   **售价提取**: 必须从指令中提取出明确的“售价”。
        *   **套餐类型提取**: 从指令中识别套餐的核心类型，例如“网费”或“包时”。将结果放入 `commodity_type` 字段。
        *   **原价计算规则**:
            *   如果 `commodity_type` 是 “**网费**”，则从指令中直接提取“原价”。例如，指令“19.9得50网费”中，“原价”是50。
            *   如果 `commodity_type` 是 “**包时**”，则“原价”应根据“售价”**估算**，规则为 **售价的3倍**。例如，指令“3小时包时，价格9.8”，`售价`是9.8，那么`原价`就应该是 9.8 * 3 = 29.4。**绝对不要**将“3小时”这个时长错误地识别为原价。
        *   **用户类型提取**: 从指令中识别用户类型，如“新客”、“新会员”应提取为 "新客"；如“老客”、“会员”应提取为 "老客"；如果未提及，则为 "不限制"。将结果放入 `member_type` 字段。
        *   **适用位置提取**: 详细分析指令中描述套餐的关键词，例如“单人双人包”、“豪华电竞包间”、“大厅”等。如果指令中明确提到了房间类型或位置，就提取它。如果未提及任何具体位置，则默认为“大厅”。将结果放入 `applicable_location` 字段。
        *   **标题生成**: 你需要根据提取出的信息，为“团购标题”生成一个清晰、规范的名称。
            *   **网费标题格式**: 保持 `【用户类型】售价得原价内容` 格式。例如：`【新客专享】42.9得100元网费`。
            *   **包时标题格式**: 简化为 `时长 + 更详细的套餐描述`。例如，指令 “包时 单人双人包套餐 5 小时，价格 39.9”，标题应生成为 `5小时单人双人包套餐`。
        *   **其他字段**: 如果指令中包含，也请提取 "可用区域", "限购", "有效期", "团单备注"。
    *   对于 `update` 操作，首先通过智能匹配找到要修改的商品。在返回结果中，`from_name` 必须使用“当前线上商品列表”中被匹配到的那个商品**完整的 `name`**。`new_data` 则是根据用户指令生成的、包含所有字段的完整新商品信息对象，其中必须包含`售价`和`原价`。
    *   对于 `delete` 操作，只需提取并返回要下架的商品的**完整的 `name`**。
4.  **格式化输出**: 必须严格按照以下JSON格式返回结果，不要添加任何额外的解释或说明文字。

```json
{{
  "add": [
    {{
      "团购标题": "...",
      "售价": 0.0,
      "原价": 0.0,
      "member_type": "新客",
      "commodity_type": "网费",
      "applicable_location": "大厅",
      "可用区域": "...",
      "限购": "...",
      "有效期": "...",
      "团单备注": "..."
    }}
  ],
  "update": [
    {{
      "from_name": "要修改的线上商品原名称",
      "new_data": {{
          "团购标题": "修改后的新名称",
          "售价": 19.9,
          "原价": 100.0,
          "可用区域": "...",
          "限购": "...",
          "有效期": "...",
          "团单备注": "..."
      }}
    }}
  ],
  "delete": [
    {{
      "name": "要下架的线上商品名称"
    }}
  ]
}}
```

**注意事项**:
-   如果找不到完全匹配的商品进行修改或下架，请不要凭空创造，在结果中忽略该项操作。
-   价格必须是数字（浮点数）。
-   返回的结果必须是纯粹的JSON字符串。
"""

    cache_key = hashlib.sha256(prompt.encode('utf-8')).hexdigest()
    if cache_key in cache:
        log_func("[Cache Hit] 发现相同分析请求的缓存结果，直接使用。")
        return cache[cache_key]

    log_func("[Cache Miss] 未找到缓存，发起新的LLM请求。")
    try:
        log_func(f"--- [DEBUG] 发送给LLM的Prompt ---\n{prompt[:800]}...\n" + "-"*30)
        response = llm_client.chat.completions.create(model=LLM_MODEL_ID, messages=[{'role': 'user', 'content': prompt}], stream=True)
        full_response = "".join(chunk.choices[0].delta.content for chunk in response if chunk.choices[0].delta.content)
        log_func(f"--- [LLM Raw Response for Text Analysis] ---\n{full_response}\n" + "-"*30)
        
        json_match = re.search(r'\{[\s\S]*\}', full_response)
        if json_match:
            cleaned_response = json_match.group(0)
            analysis_result = json.loads(cleaned_response)
        else:
            raise json.JSONDecodeError("在LLM响应中未找到JSON对象", full_response, 0)
        
        log_func("[Success] LLM文本指令分析成功！")
        cache[cache_key] = analysis_result
        log_func(f"已将本次结果存入缓存，Key: {cache_key[:10]}...")
        return analysis_result
    except Exception as e:
        log_func(f"[Error] LLM文本指令分析过程中出错: {e}\n{traceback.format_exc()}")
    return None

def parse_product_details(details):
    """Parses the raw product detail JSON into a clean dictionary."""
    if not details or 'product' not in details:
        raise ValueError("无效的商品详情数据")

    product = details.get('product', {})
    sku = details.get('skus', [{}])[0]
    attr_map = product.get('attr_key_value_map', {})

    name = product.get('product_name')
    price = sku.get('actual_amount', 0) / 100
    product_id = product.get('product_id')

    # Default values
    area, limit, validity, notes = "未知", "未知", "未知", ""

    try:
        notification = json.loads(attr_map.get('Notification', '[]'))
        title_map = {item['title']: item['content'] for item in notification}
        
        validity_text = title_map.get('有效期', '购买后30日内有效')
        validity = validity_text.replace("购买后", "").replace("内有效", "")
        
        limit = title_map.get('限购说明', '无')
        notes = title_map.get('使用须知', '')
        
        desc = json.loads(attr_map.get('Description', '[]'))
        if desc:
            area = desc[0].replace("适用区域: ", "")
    except (json.JSONDecodeError, IndexError, KeyError) as e:
        # Silently ignore parsing errors, use defaults
        pass

    return {
        "id": product_id,
        "团购标题": name,
        "售价": price,
        "可用区域": area,
        "限购": limit,
        "有效期": validity,
        "团单备注": notes
    }

def center_crop_image(img, aspect_ratio):
    """居中裁剪图片"""
    width, height = img.size
    target_width, target_height = width, width / aspect_ratio
    if target_height > height:
        target_height = height
        target_width = height * aspect_ratio
    left, top = (width - target_width) / 2, (height - target_height) / 2
    return img.crop((left, top, left + target_width, top + target_height))

def upload_to_r2(img_obj, poi_id, aspect_ratio_str, log_func):
    """上传图片到Cloudflare R2"""
    log_func(f"--- 正在上传 {aspect_ratio_str} 比例的图片 ---")
    try:
        s3_client = boto3.client('s3', endpoint_url=R2_ENDPOINT_URL, aws_access_key_id=CLOUDFLARE_R2_ACCESS_KEY_ID, aws_secret_access_key=CLOUDFLARE_R2_SECRET_ACCESS_KEY, config=Config(signature_version='s3v4'))
        in_mem_file = BytesIO()
        if img_obj.mode == 'RGBA':
            img_obj = img_obj.convert('RGB')
        img_obj.save(in_mem_file, format='JPEG', quality=90)
        in_mem_file.seek(0)
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{poi_id}_{timestamp}_{aspect_ratio_str.replace(':', '_')}.jpg"
        s3_client.upload_fileobj(in_mem_file, R2_BUCKET_NAME, filename, ExtraArgs={'ContentType': 'image/jpeg'})
        image_url = f"{R2_PUBLIC_URL_PREFIX}/{filename}"
        log_func(f"[Success] 图片上传成功: {image_url}")
        return image_url
    except Exception as e:
        log_func(f"[Error] 图片上传到R2失败: {e}")
        return None

class App(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("抖音团购智能同步工具 v2.4")
        self.master.geometry("1350x800")
        self.pack(fill="both", expand=True)
        self.store_data, self.douyin_access_token, self.douyin_products, self.excel_data, self.excel_file_path, self.all_store_names, self.image_dir, self.current_poi_id = {}, None, [], [], "", [], None, None
        self.llm_cache = {}
        self.product_details_cache = {}
        self.hide_live_only_var = tk.BooleanVar(value=False)
        self.multi_match_mode_var = tk.BooleanVar(value=False)
        self.edit_entry = None
        self.create_widgets()
        self.init_backend()

    def log(self, message): self.master.after(0, lambda: self._log_thread_safe(message))
    def _log_thread_safe(self, message):
        # GUI日志
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        # 文件日志
        logging.info(message)

    def create_widgets(self):
        top_frame = ttk.Frame(self); top_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(top_frame, text="选择门店:").pack(side="left", padx=(0, 5))
        self.store_name_combobox = ttk.Combobox(top_frame, width=35); self.store_name_combobox.pack(side="left", padx=5)
        self.store_name_combobox.bind('<KeyRelease>', self.filter_combobox_list)
        
        ttk.Label(top_frame, text="排除关键词:").pack(side="left", padx=(10, 5))
        self.exclude_keyword_var = tk.StringVar()
        self.exclude_keyword_entry = ttk.Entry(top_frame, textvariable=self.exclude_keyword_var, width=20); self.exclude_keyword_entry.pack(side="left")

        self.query_douyin_btn = ttk.Button(top_frame, text="1. 查询抖音商品", command=self.start_query_douyin); self.query_douyin_btn.pack(side="left", padx=5)
        self.hide_live_only_check = ttk.Checkbutton(top_frame, text="隐藏仅直播间商品", variable=self.hide_live_only_var, command=self.filter_and_populate_products); self.hide_live_only_check.pack(side="left", padx=10)
        self.web_config_btn = ttk.Button(top_frame, text="⚙️ 网页端配置", command=self.open_web_config); self.web_config_btn.pack(side="left", padx=5)

        excel_frame = ttk.LabelFrame(self, text="Excel 数据源"); excel_frame.pack(fill="x", padx=10, pady=5)
        self.load_excel_btn = ttk.Button(excel_frame, text="2. 加载Excel文件", command=self.start_load_excel); self.load_excel_btn.pack(side="left", padx=5, pady=5)
        self.excel_path_label = ttk.Label(excel_frame, text="未加载文件"); self.excel_path_label.pack(side="left", padx=5)

        image_source_frame = ttk.LabelFrame(self, text="图片源 (用于新增套餐)"); image_source_frame.pack(fill="x", padx=10, pady=5)
        self.select_image_dir_btn = ttk.Button(image_source_frame, text="选择图片文件夹", command=self.select_image_dir); self.select_image_dir_btn.pack(side="left", padx=5, pady=5)
        self.image_dir_label = ttk.Label(image_source_frame, text="未选择文件夹"); self.image_dir_label.pack(side="left", padx=5)
        self.auto_match_image_btn = ttk.Button(image_source_frame, text="自动匹配图片", command=self.start_auto_match_images, state="disabled"); self.auto_match_image_btn.pack(side="left", padx=10)
        self.multi_match_check = ttk.Checkbutton(image_source_frame, text="图片匹配多套餐模式", variable=self.multi_match_mode_var); self.multi_match_check.pack(side="left", padx=10)

        analysis_frame = ttk.LabelFrame(self, text="智能分析 (文本/图片/美团同步)"); analysis_frame.pack(fill="x", padx=10, pady=5)
        self.analysis_text = tk.Text(analysis_frame, height=8, width=80); self.analysis_text.pack(side="left", fill="x", expand=True, padx=5, pady=5)
        analysis_btn_frame = ttk.Frame(analysis_frame); analysis_btn_frame.pack(side="left", fill="y", padx=5)
        self.analyze_text_btn = ttk.Button(analysis_btn_frame, text="分析文本", command=self.start_text_analysis); self.analyze_text_btn.pack(pady=5, fill="x")
        self.sync_meituan_btn = ttk.Button(analysis_btn_frame, text="同步美团套餐", command=self.start_sync_meituan); self.sync_meituan_btn.pack(pady=5, fill="x")
        # 美团同步选项
        self.meituan_skip_update_var = tk.BooleanVar(value=False)
        self.meituan_skip_update_check = ttk.Checkbutton(analysis_btn_frame, text="仅新增/下架\n(跳过价格更新)", variable=self.meituan_skip_update_var); self.meituan_skip_update_check.pack(pady=2, fill="x")
        self.analyze_image_btn = ttk.Button(analysis_btn_frame, text="分析图片 (暂未实现)", state="disabled"); self.analyze_image_btn.pack(pady=5, fill="x")

        main_frame = ttk.Frame(self); main_frame.pack(fill="both", expand=True, padx=10, pady=5)
        columns = ("douyin_name", "douyin_price", "douyin_origin_price", "match_status", "action_mode", "matched_image", "excel_title", "excel_price", "excel_origin_price", "commodity_type", "applicable_location", "excel_area", "excel_limit", "excel_validity", "id")
        self.product_tree = ttk.Treeview(main_frame, columns=columns, show="headings"); self.product_tree.pack(side="left", fill="both", expand=True)
        headings = {
            "douyin_name": "抖音商品名", "douyin_price": "抖音价", "douyin_origin_price": "抖音原价",
            "match_status": "匹配状态", "action_mode": "操作模式", "matched_image": "匹配图片",
            "excel_title": "匹配Excel商品", "excel_price": "Excel价", "excel_origin_price": "Excel原价",
            "commodity_type": "套餐类型", "applicable_location": "适用位置", "excel_area": "可用区域", "excel_limit": "限购", "excel_validity": "有效期",
            "id": "Product ID"
        }
        widths = {
            "douyin_name": 220, "douyin_price": 60, "douyin_origin_price": 60,
            "match_status": 80, "action_mode": 80, "matched_image": 120,
            "excel_title": 220, "excel_price": 60, "excel_origin_price": 60,
            "commodity_type": 80, "applicable_location": 100, "excel_area": 100, "excel_limit": 80, "excel_validity": 60
        }
        for col, text in headings.items(): self.product_tree.heading(col, text=text)
        for col, width in widths.items(): self.product_tree.column(col, width=width, anchor="center" if "price" in col or "status" in col or "validity" in col or "mode" in col else "w")
        self.product_tree["displaycolumns"] = [col for col in columns if col != "id"]
        self.product_tree.bind("<Double-1>", self.edit_cell)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.product_tree.yview); self.product_tree.configure(yscroll=scrollbar.set); scrollbar.pack(side="right", fill="y")
        
        bottom_frame = ttk.Frame(self); bottom_frame.pack(fill="both", expand=True, padx=10, pady=5)
        action_frame = ttk.Frame(bottom_frame); action_frame.pack(fill="x", pady=5)
        self.match_btn = ttk.Button(action_frame, text="3. 开始智能匹配", command=self.start_match_products, state="disabled"); self.match_btn.pack(side="left", fill="x", expand=True, padx=5)
        self.delete_btn = ttk.Button(action_frame, text="删除选中行", command=self.delete_selected_rows); self.delete_btn.pack(side="left", fill="x", expand=True, padx=5)
        self.update_btn = ttk.Button(action_frame, text="4. 一键批量操作", command=self.start_batch_update, state="disabled"); self.update_btn.pack(side="left", fill="x", expand=True, padx=5)
        log_frame = ttk.LabelFrame(bottom_frame, text="日志"); log_frame.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_frame, wrap="word", state="disabled", height=10); self.log_text.pack(fill="both", expand=True)

    def set_ui_state(self, is_busy):
        state = "disabled" if is_busy else "normal"
        for btn in (self.query_douyin_btn, self.load_excel_btn, self.delete_btn, self.analyze_text_btn, self.select_image_dir_btn, self.auto_match_image_btn): btn.config(state=state)
        self.store_name_combobox.config(state="disabled" if is_busy else "normal")
        self.exclude_keyword_entry.config(state="disabled" if is_busy else "normal")
        self.match_btn.config(state="disabled" if is_busy or not (self.douyin_products and self.excel_data) else "normal")
        self.update_btn.config(state="disabled" if is_busy or not any(self.product_tree.set(item, "match_status") == "匹配成功" for item in self.product_tree.get_children()) else "normal")
        self.auto_match_image_btn.config(state="disabled" if is_busy or not self.image_dir else "normal")

    def open_web_config(self):
        """打开网页端配置对话框"""
        config_window = tk.Toplevel(self.master)
        config_window.title("网页端配置（用于重创模式）")
        config_window.geometry("750x500")
        
        # 标题
        ttk.Label(config_window, text="重创模式 - 网页端API配置", font=("", 12, "bold")).pack(pady=10)
        
        # 配置状态
        status_frame = ttk.LabelFrame(config_window, text="当前配置状态")
        status_frame.pack(fill="x", padx=20, pady=10)
        
        cookie_status = "✅ 已配置" if DOUYIN_WEB_COOKIE else "❌ 未配置"
        csrf_status = "✅ 已配置" if DOUYIN_WEB_CSRF_TOKEN else "❌ 未配置"
        
        ttk.Label(status_frame, text=f"Cookie: {cookie_status}", font=("", 10)).pack(anchor="w", padx=10, pady=5)
        ttk.Label(status_frame, text=f"CSRF Token: {csrf_status}", font=("", 10)).pack(anchor="w", padx=10, pady=5)
        ttk.Label(status_frame, text=f"Cookie来源: cookie.txt 文件", font=("", 9), foreground="gray").pack(anchor="w", padx=10, pady=2)
        ttk.Label(status_frame, text=f"CSRF Token: 硬编码", font=("", 9), foreground="gray").pack(anchor="w", padx=10, pady=2)
        
        # 说明文字
        info_frame = ttk.LabelFrame(config_window, text="配置说明")
        info_frame.pack(fill="x", padx=20, pady=10)
        
        help_text = """
Cookie配置：
• Cookie已从程序根目录的 cookie.txt 文件自动加载
• 如需更新Cookie，请编辑 cookie.txt 文件，然后点击下方"重新加载Cookie"按钮

CSRF Token配置：
• CSRF Token已硬编码在程序中
• 当前值: 000100000001ae8a406b9344d0cc4e30ceaf542c505dbbabca5a3842c450a93e0787a4d2f8991880c8ea9d2d1372

如何获取新的Cookie：
1. 打开浏览器，登录 https://life.douyin.com
2. 按F12打开开发者工具，切换到Network标签
3. 刷新页面，找到任意请求
4. 在Request Headers中复制完整的Cookie值
5. 将Cookie粘贴到程序根目录的 cookie.txt 文件中
        """
        ttk.Label(info_frame, text=help_text, justify="left", foreground="#333").pack(anchor="w", padx=10, pady=10)
        
        # 按钮区域
        button_frame = ttk.Frame(config_window)
        button_frame.pack(pady=10)
        
        def reload_cookie():
            global DOUYIN_WEB_COOKIE
            DOUYIN_WEB_COOKIE = load_cookie_from_file()
            if DOUYIN_WEB_COOKIE:
                messagebox.showinfo("重新加载成功", "Cookie已从 cookie.txt 文件重新加载！")
                config_window.destroy()
                self.open_web_config()  # 重新打开窗口显示新状态
            else:
                messagebox.showerror("加载失败", "无法从 cookie.txt 文件加载Cookie，请检查文件是否存在。")
        
        def open_cookie_file():
            cookie_file = os.path.join(os.path.dirname(__file__), 'cookie.txt')
            if os.path.exists(cookie_file):
                os.startfile(cookie_file)
            else:
                messagebox.showwarning("文件不存在", f"Cookie文件不存在: {cookie_file}\n\n请创建该文件并粘贴Cookie内容。")
        
        ttk.Button(button_frame, text="🔄 重新加载Cookie", command=reload_cookie).pack(side="left", padx=5)
        ttk.Button(button_frame, text="📝 打开cookie.txt", command=open_cookie_file).pack(side="left", padx=5)
        ttk.Button(button_frame, text="关闭", command=config_window.destroy).pack(side="left", padx=5)

    def init_backend(self):
        self.set_ui_state(True)
        threading.Thread(target=self._init_backend_thread, daemon=True).start()

    def _init_backend_thread(self):
        self.douyin_access_token = get_douyin_access_token(self.log)
        if not self.douyin_access_token: messagebox.showerror("错误", "获取抖音Access Token失败。")
        feishu_token = get_feishu_tenant_access_token(self.log)
        if feishu_token:
            self.store_data = get_feishu_bitable_records(feishu_token, self.log)
            if self.store_data: self.master.after(0, self.update_store_combobox)
        self.master.after(0, lambda: self.set_ui_state(False))

    def select_image_dir(self):
        dir_path = filedialog.askdirectory(title="选择包含头图的文件夹")
        if dir_path:
            self.image_dir = dir_path
            self.image_dir_label.config(text=f"已选: {os.path.basename(dir_path)}")
            self.log(f"新增套餐的图片源文件夹已设置为: {dir_path}")

    def update_store_combobox(self):
        self.all_store_names = sorted(list(self.store_data.keys()))
        self.store_name_combobox['values'] = self.all_store_names
        if self.all_store_names: self.store_name_combobox.set(self.all_store_names[0])
        self.log("门店下拉列表已更新。")

    def filter_combobox_list(self, event):
        typed_text = self.store_name_combobox.get().lower()
        self.store_name_combobox['values'] = [name for name in self.all_store_names if typed_text in name.lower()] if typed_text else self.all_store_names

    def start_query_douyin(self):
        store_name = self.store_name_combobox.get().strip()
        if not store_name: return messagebox.showerror("错误", "请选择一个门店。")
        poi_id = self.store_data.get(store_name)
        if not poi_id: return messagebox.showerror("错误", f"未找到门店 '{store_name}' 的ID。")
        self.set_ui_state(True)
        threading.Thread(target=self._query_douyin_thread, args=(poi_id,), daemon=True).start()
    
    def _query_douyin_thread(self, poi_id):
        self.current_poi_id = poi_id
        self.product_details_cache.clear()
        all_products = get_douyin_products_by_store(self.douyin_access_token, poi_id, self.log)
        
        exclude_string = self.exclude_keyword_var.get().strip()
        if exclude_string:
            # Use regex to split by spaces, commas (English/Chinese), and semicolons
            exclude_keywords = [kw for kw in re.split(r'[ ,;，；]+', exclude_string) if kw]
            if exclude_keywords:
                self.log(f"开始根据排除关键词 {exclude_keywords} 过滤商品...")
                # Filter products: exclude if ANY keyword is present in the name
                self.douyin_products = [
                    p for p in all_products
                    if not any(keyword in p['name'] for keyword in exclude_keywords)
                ]
                self.log(f"过滤后剩余 {len(self.douyin_products)} 个商品。")
            else:
                self.douyin_products = all_products
        else:
            self.douyin_products = all_products

        self.master.after(0, self.filter_and_populate_products)
        self.master.after(0, lambda: self.set_ui_state(False))

    def populate_product_list(self, products_to_show):
        self.product_tree.delete(*self.product_tree.get_children())
        for pkg in products_to_show:
            # columns = ("douyin_name", "douyin_price", "douyin_origin_price", "match_status", "action_mode", "excel_title", "excel_price", "excel_origin_price", "excel_area", "excel_limit", "excel_validity", "id")
            values = (
                pkg['name'], pkg['price'], pkg.get('origin_price', '0.00'),
                "未匹配", "修改", "",
                "", "", "", "", "", "",
                pkg['id']
            )
            self.product_tree.insert("", "end", values=values)
        self.log(f"抖音商品列表已更新，共 {len(products_to_show)} 项。")
    
    def filter_and_populate_products(self):
        self.set_ui_state(True)
        self.log("正在应用筛选条件...")
        threading.Thread(target=self._filter_worker_thread, daemon=True).start()

    def _filter_worker_thread(self):
        hide_live_only = self.hide_live_only_var.get()
        if not hide_live_only:
            self.master.after(0, lambda: self.populate_product_list(self.douyin_products))
            self.master.after(0, lambda: self.set_ui_state(False))
            return

        self.log("筛选开启：隐藏仅直播间可见商品。这可能需要一些时间...")
        filtered_products = []
        total = len(self.douyin_products)
        for i, p in enumerate(self.douyin_products):
            product_id = p['id']
            
            if product_id not in self.product_details_cache:
                self.log(f"正在获取商品详情 ({i+1}/{total}): {p['name'][:20]}...")
                details = get_douyin_product_details(self.douyin_access_token, product_id, self.log)
                self.product_details_cache[product_id] = details
            else:
                details = self.product_details_cache[product_id]
            
            if not details or 'product' not in details:
                self.log(f"[Warning] 无法获取 {product_id} 的详情，将隐藏该商品。")
                continue

            attr_map = details.get('product', {}).get('attr_key_value_map', {})
            show_channel = str(attr_map.get('show_channel', '1'))
            
            if show_channel == '2':
                self.log(f" -> 已隐藏 (仅直播间): {p['name']}")
                continue
            
            filtered_products.append(p)

        self.master.after(0, lambda: self.populate_product_list(filtered_products))
        self.master.after(0, lambda: self.set_ui_state(False))

    def start_load_excel(self):
        file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        if not file_path: return
        self.excel_file_path = file_path; self.set_ui_state(True)
        threading.Thread(target=self._load_excel_thread, args=(file_path,), daemon=True).start()

    def _load_excel_thread(self, file_path):
        self.excel_data = load_excel_data(file_path, self.log)
        if self.excel_data is not None: self.master.after(0, lambda: self.excel_path_label.config(text=os.path.basename(file_path)))
        else: messagebox.showerror("错误", "加载Excel文件失败。")
        self.master.after(0, lambda: self.set_ui_state(False))

    def select_image_dir(self):
        dir_path = filedialog.askdirectory(title="选择包含头图的文件夹")
        if dir_path:
            self.image_dir = dir_path
            self.image_dir_label.config(text=f"已选: {os.path.basename(dir_path)}")
            self.log(f"新增套餐的图片源文件夹已设置为: {dir_path}")

    def start_match_products(self):
        if not self.excel_data:
            messagebox.showinfo("提示", "请先加载Excel文件。")
            return

        current_products_in_tree = []
        for item_id in self.product_tree.get_children():
            douyin_name = self.product_tree.set(item_id, "douyin_name")
            # 只对实际的抖音商品进行匹配，忽略待创建的行
            if douyin_name and douyin_name != "<待创建>":
                current_products_in_tree.append({
                    "name": douyin_name,
                    "price": self.product_tree.set(item_id, "douyin_price"),
                    "origin_price": self.product_tree.set(item_id, "douyin_origin_price"),
                    "id": self.product_tree.set(item_id, "id")
                })
        
        if not current_products_in_tree:
            messagebox.showinfo("提示", "抖音商品列表为空，无法进行匹配。")
            return

        self.set_ui_state(True)
        threading.Thread(target=self._match_products_thread, args=(current_products_in_tree,), daemon=True).start()

    def _match_products_thread(self, current_douyin_products):
        match_result = match_products_with_llm(current_douyin_products, self.excel_data, self.log, self.llm_cache)
        if match_result: self.master.after(0, lambda: self.update_matches_in_tree(match_result))
        self.master.after(0, lambda: self.set_ui_state(False))

    def update_matches_in_tree(self, match_result):
        excel_map = {item['团购标题']: item for item in self.excel_data}
        for item_id in self.product_tree.get_children():
            douyin_name = self.product_tree.set(item_id, "douyin_name")
            matched_excel_title = match_result.get(douyin_name)
            if matched_excel_title in excel_map:
                excel_item = excel_map[matched_excel_title]
                self.product_tree.set(item_id, "match_status", "匹配成功")
                self.product_tree.set(item_id, "excel_title", matched_excel_title)
                self.product_tree.set(item_id, "excel_price", excel_item.get('售价', ''))
                self.product_tree.set(item_id, "excel_area", excel_item.get('可用区域', ''))
                self.product_tree.set(item_id, "excel_limit", excel_item.get('限购', ''))
                self.product_tree.set(item_id, "excel_validity", excel_item.get('有效期', ''))
                self.product_tree.set(item_id, "action_mode", "重创" if "重置次数" in str(excel_item.get('可用区域', '')) else "修改")
            else:
                self.product_tree.set(item_id, "match_status", "匹配失败")
                self.product_tree.set(item_id, "action_mode", "-")
                for col in ("excel_title", "excel_price", "excel_area", "excel_limit", "excel_validity"): self.product_tree.set(item_id, col, "")
        self.log("LLM匹配结果已更新到界面。")

    def edit_cell(self, event):
        if hasattr(self, 'edit_entry') and self.edit_entry:
            self.edit_entry.destroy()

        item_id = self.product_tree.focus()
        if not item_id: return

        column_id = self.product_tree.identify_column(event.x)
        column_name = self.product_tree.column(column_id, "id")

        if column_name == "action_mode":
            current_mode = self.product_tree.set(item_id, "action_mode")
            modes = ["修改", "重创", "下架", "-"]
            try:
                new_mode = modes[(modes.index(current_mode) + 1) % len(modes)]
            except ValueError:
                new_mode = "修改"
            self.product_tree.set(item_id, "action_mode", new_mode)
            return

        editable_columns = ["excel_title", "excel_price", "excel_origin_price", "commodity_type", "applicable_location", "excel_area", "excel_limit", "excel_validity", "matched_image"]
        if column_name not in editable_columns:
            return

        x, y, width, height = self.product_tree.bbox(item_id, column_id)
        value = self.product_tree.set(item_id, column_name)
        
        entry_var = tk.StringVar(value=value)
        self.edit_entry = ttk.Entry(self.product_tree, textvariable=entry_var)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.focus_force()
        self.edit_entry.selection_range(0, tk.END)

        def on_edit_done(event=None):
            new_value = entry_var.get()
            self.product_tree.set(item_id, column_name, new_value)
            if hasattr(self, 'edit_entry') and self.edit_entry:
                self.edit_entry.destroy()
                self.edit_entry = None
        
        self.edit_entry.bind("<Return>", on_edit_done)
        self.edit_entry.bind("<FocusOut>", on_edit_done)
        self.edit_entry.bind("<Escape>", lambda e: self.edit_entry.destroy() if hasattr(self, 'edit_entry') and self.edit_entry else None)

    def delete_selected_rows(self):
        selected_items = self.product_tree.selection()
        if not selected_items: return messagebox.showinfo("提示", "请先选择要删除的行。")
        if messagebox.askyesno("确认删除", f"确定要删除选中的 {len(selected_items)} 行吗？此操作仅在界面上移除，不影响线上商品。"):
            for item in selected_items: self.product_tree.delete(item)
            self.log(f"已从界面删除 {len(selected_items)} 行。")

    def start_sync_meituan(self):
        """开始美团同步"""
        if not self.douyin_products:
            messagebox.showerror("错误", "请先查询抖音商品列表")
            return
        
        store_name = self.store_name_combobox.get().strip()
        if not store_name:
            messagebox.showerror("错误", "请选择门店")
            return
        
        # 使用simpledialog询问城市拼音
        from tkinter import simpledialog
        city = simpledialog.askstring(
            "输入城市信息",
            "请输入城市拼音（如：taiyuan）:",
            initialvalue="taiyuan",
            parent=self.master
        )
        
        if not city or not city.strip():
            return
        
        self.set_ui_state(True)
        threading.Thread(target=self._sync_meituan_thread, args=(store_name, city.strip()), daemon=True).start()
    
    def _sync_meituan_thread(self, store_name, city):
        """美团同步线程"""
        self.log("========== 开始美团同步 ==========")
        
        # 1. 处理店名
        cleaned_store_name = process_store_name_for_meituan(store_name, self.log)
        
        # 2. 获取美团套餐
        meituan_packages = get_meituan_packages(cleaned_store_name, city, self.log)
        
        if not meituan_packages:
            self.log("[Error] 未能获取美团套餐，同步终止")
            self.master.after(0, lambda: messagebox.showerror("错误", "未能获取美团套餐，请检查网络和代理设置"))
            self.master.after(0, lambda: self.set_ui_state(False))
            return
        
        # 3. 智能匹配
        match_result = match_packages_smart(self.douyin_products, meituan_packages, self.log)
        
        # 获取用户选项：是否跳过价格更新
        skip_price_update = self.meituan_skip_update_var.get()
        if skip_price_update:
            self.log("\n[用户选项] 已启用'仅新增/下架'模式，将跳过所有价格更新操作")
        
        # 4. 构建操作列表
        operations = []
        
        # 4.1 匹配的套餐 - 根据action和用户选项决定是否更新
        for match in match_result["matches"]:
            dy_pkg = match["douyin"]
            mt_pkg = match["meituan"]
            action = match["action"]
            
            if action == "update":
                if skip_price_update:
                    # 用户选择跳过价格更新
                    self.log(f"跳过价格更新: {dy_pkg['name']} (用户选择仅新增/下架)")
                else:
                    # 价格不同，需要更新
                    operations.append({
                        "action": "update",
                        "product_id": dy_pkg['id'],
                        "douyin_name": dy_pkg['name'],
                        "new_data": {
                            "团购标题": dy_pkg['name'],  # 保持原名称
                            "售价": mt_pkg['price'],
                            "原价": mt_pkg['original_price'],
                            "可用区域": "",
                            "限购": "",
                            "有效期": "",
                            "团单备注": ""
                        }
                    })
            elif action == "keep":
                # 价格相同，保持原样，不添加到操作列表
                self.log(f"保持原样: {dy_pkg['name']} (价格已同步)")
                pass
        
        # 4.2 美团独有 - 新建
        for mt_pkg in match_result["meituan_only"]:
            operations.append({
                "action": "add",
                "product_id": None,
                "douyin_name": "<待创建>",
                "new_data": {
                    "团购标题": mt_pkg['title'],
                    "售价": mt_pkg['price'],
                    "原价": mt_pkg['original_price'],
                    "member_type": "不限制",
                    "commodity_type": "网费" if "网费" in mt_pkg['title'] else "包时",
                    "applicable_location": "大厅",
                    "可用区域": "",
                    "限购": "",
                    "有效期": "30",
                    "团单备注": ""
                }
            })
        
        # 4.3 抖音独有 - 下架
        for dy_pkg in match_result["douyin_only"]:
            operations.append({
                "action": "delete",
                "product_id": dy_pkg['id'],
                "douyin_name": dy_pkg['name'],
                "new_data": {"团购标题": f"下架-{dy_pkg['name']}"}
            })
        
        # 5. 统计操作类型
        update_count = sum(1 for op in operations if op["action"] == "update")
        add_count = sum(1 for op in operations if op["action"] == "add")
        delete_count = sum(1 for op in operations if op["action"] == "delete")
        
        # 6. 更新UI显示
        self.master.after(0, lambda: self._populate_meituan_sync_result(operations))
        self.master.after(0, lambda: self.set_ui_state(False))
        
        # 7. 输出汇总信息
        self.log(f"\n========== 美团同步分析完成 ==========")
        self.log(f"总操作数: {len(operations)} 个")
        self.log(f"  - 价格更新: {update_count} 个")
        self.log(f"  - 新增套餐: {add_count} 个")
        self.log(f"  - 下架套餐: {delete_count} 个")
        if skip_price_update and update_count == 0:
            skipped_updates = sum(1 for m in match_result["matches"] if m["action"] == "update")
            if skipped_updates > 0:
                self.log(f"  - 已跳过价格更新: {skipped_updates} 个（用户选择）")
        self.log(f"========================================\n")
    
    def _populate_meituan_sync_result(self, operations):
        """将美团同步结果填充到表格"""
        self.product_tree.delete(*self.product_tree.get_children())
        self.excel_data = []
        
        for op in operations:
            action = op["action"]
            product_id = op.get("product_id", "")
            douyin_name = op.get("douyin_name", "")
            new_data = op["new_data"]
            
            self.excel_data.append(new_data)
            
            if action == "update":
                action_mode = "修改"
                douyin_price = next((p['price'] for p in self.douyin_products if p['id'] == product_id), "")
                douyin_origin_price = next((p['origin_price'] for p in self.douyin_products if p['id'] == product_id), "")
            elif action == "add":
                action_mode = "重创"
                douyin_price = ""
                douyin_origin_price = ""
            else:  # delete
                action_mode = "下架"
                douyin_price = next((p['price'] for p in self.douyin_products if p['id'] == product_id), "")
                douyin_origin_price = next((p['origin_price'] for p in self.douyin_products if p['id'] == product_id), "")
            
            values = (
                douyin_name, douyin_price, douyin_origin_price,
                "匹配成功", action_mode, "",
                new_data.get('团购标题'), new_data.get('售价'), new_data.get('原价'),
                new_data.get('commodity_type', ''), new_data.get('applicable_location', ''),
                new_data.get('可用区域'), new_data.get('限购'), new_data.get('有效期'),
                product_id
            )
            
            tag = 'update' if action == "update" else ('add' if action == "add" else 'delete')
            self.product_tree.insert("", "end", values=values, tags=(tag,))
        
        self.product_tree.tag_configure('add', background='#D4EDDA')
        self.product_tree.tag_configure('update', background='#FFF3CD')
        self.product_tree.tag_configure('delete', background='#F8D7DA')
        
        self.log("美团同步结果已更新到界面，请检查后执行'一键批量操作'")
        self.update_btn.config(state="normal")

    def start_text_analysis(self):
        if not self.douyin_products:
            if not messagebox.askyesno("确认操作", "当前门店没有线上商品或未查询。\n\n这会导致AI无法进行'修改'或'下架'的判断。\n\n是否继续，只执行纯'新增'操作？"):
                return
        
        text_to_analyze = self.analysis_text.get("1.0", tk.END).strip()
        if not text_to_analyze:
            messagebox.showerror("错误", "请输入需要分析的文本内容。")
            return
            
        if not messagebox.askyesno("确认操作", "这将使用LLM分析文本并覆盖当前表格内容，确定要继续吗？"):
            return

        self.set_ui_state(True)
        threading.Thread(target=self._text_analysis_thread, args=(text_to_analyze,), daemon=True).start()

    def _text_analysis_thread(self, text_to_analyze):
        self.log("--- 开始使用LLM进行文本智能分析 ---")
        
        # 直接使用已有的简略商品列表，无需重新获取详情
        simple_product_list = self.douyin_products
        self.log(f"已加载 {len(simple_product_list)} 个线上商品用于分析。")

        analysis_result = analyze_text_for_actions(text_to_analyze, simple_product_list, self.log, self.llm_cache)
        
        if analysis_result:
            self.master.after(0, lambda: self.populate_tree_from_analysis(analysis_result))
        else:
            self.log("[Error] 文本分析失败，未能获取有效结果。")
            
        self.master.after(0, lambda: self.set_ui_state(False))

    def populate_tree_from_analysis(self, analysis_result):
        self.log("--- 正在根据文本分析结果更新列表 ---")
        self.product_tree.delete(*self.product_tree.get_children())
        
        douyin_products_map = {p['name']: p for p in self.douyin_products}
        self.excel_data = []
        processed_douyin_products = set()

        # columns = ("douyin_name", "douyin_price", "douyin_origin_price", "match_status", "action_mode", "excel_title", "excel_price", "excel_origin_price", "excel_area", "excel_limit", "excel_validity", "id")
        for new_data in analysis_result.get('add', []):
            self.excel_data.append(new_data)
            values = (
                "<待创建>", "", "",
                "匹配成功", "重创", "",
                new_data.get('团购标题'), new_data.get('售价'), new_data.get('原价'),
                new_data.get('commodity_type'), new_data.get('applicable_location', '大厅'), new_data.get('可用区域'), new_data.get('限购'), new_data.get('有效期'), ""
            )
            self.product_tree.insert("", "end", values=values, tags=('add',))

        for update_item in analysis_result.get('update', []):
            from_name = update_item.get('from_name')
            new_data = update_item.get('new_data')
            if from_name in douyin_products_map and new_data:
                product = douyin_products_map[from_name]
                self.excel_data.append(new_data)
                values = (
                    product['name'], product['price'], product.get('origin_price', '0.00'),
                    "匹配成功", "修改", "",
                    new_data.get('团购标题'), new_data.get('售价'), new_data.get('原价'),
                    new_data.get('commodity_type', ''), new_data.get('applicable_location', ''),
                    new_data.get('可用区域'), new_data.get('限购'), new_data.get('有效期'), product['id']
                )
                self.product_tree.insert("", "end", values=values, tags=('update',))
                processed_douyin_products.add(from_name)

        for delete_item in analysis_result.get('delete', []):
            name = delete_item.get('name')
            if name in douyin_products_map:
                product = douyin_products_map[name]
                new_data = {"团购标题": f"下架-{name}"}
                self.excel_data.append(new_data)
                values = (
                    product['name'], product['price'], product.get('origin_price', '0.00'),
                    "匹配成功", "下架", "",
                    f"下架-{name}", "", "", "", "", "", product['id']
                )
                self.product_tree.insert("", "end", values=values, tags=('delete',))
                processed_douyin_products.add(name)

        for name, product in douyin_products_map.items():
            if name not in processed_douyin_products:
                values = (
                    product['name'], product['price'], product.get('origin_price', '0.00'),
                    "无操作", "-", "", "", "", "", "", "", "", product['id']
                )
                self.product_tree.insert("", "end", values=values, tags=('keep',))
        
        self.product_tree.tag_configure('add', background='#D4EDDA')
        self.product_tree.tag_configure('update', background='#FFF3CD')
        self.product_tree.tag_configure('delete', background='#F8D7DA')

        self.log("列表已根据分析结果更新。请检查并执行“一键批量操作”。")
        self.update_btn.config(state="normal")

    def start_batch_update(self):
        if hasattr(self, 'edit_entry') and self.edit_entry:
            self.edit_entry.destroy()
            self.edit_entry = None
            
        items_to_process = []
        for item_id in self.product_tree.get_children():
            action_mode = self.product_tree.set(item_id, "action_mode")
            if action_mode == "-" or action_mode == "无操作":
                continue

            try:
                # 从Treeview中直接读取最终确认的数据
                excel_title_from_tree = self.product_tree.set(item_id, "excel_title")
                full_data_from_excel = next((d for d in self.excel_data if d.get("团购标题") == excel_title_from_tree), {})

                new_data = {
                    "团购标题": excel_title_from_tree,
                    "售价": float(self.product_tree.set(item_id, "excel_price") or 0),
                    "原价": float(self.product_tree.set(item_id, "excel_origin_price") or 0),
                    "可用区域": self.product_tree.set(item_id, "excel_area"),
                    "限购": self.product_tree.set(item_id, "excel_limit"),
                    "有效期": self.product_tree.set(item_id, "excel_validity"),
                    "团单备注": "", # 备注字段目前不在表格中，默认为空
                    "matched_image": self.product_tree.set(item_id, "matched_image"),
                    "member_type": full_data_from_excel.get("member_type"),
                    "commodity_type": self.product_tree.set(item_id, "commodity_type"),
                    "applicable_location": self.product_tree.set(item_id, "applicable_location")
                }
                
                # 对“下架”操作进行特殊处理
                if action_mode == "下架":
                    original_name = self.product_tree.set(item_id, "douyin_name")
                    new_data["团购标题"] = f"下架-{original_name}"

                item_to_add = {
                    "product_id": self.product_tree.set(item_id, "id"),
                    "new_data": new_data,
                    "action_mode": action_mode
                }
                items_to_process.append(item_to_add)
            except ValueError:
                messagebox.showerror("数据错误", f"商品 '{self.product_tree.set(item_id, 'excel_title')}' 的价格格式不正确，请确保为数字。")
                return
            except Exception as e:
                messagebox.showerror("未知错误", f"处理行数据时发生错误: {e}")
                return

        if not items_to_process: return messagebox.showinfo("提示", "没有找到任何需要操作的商品（操作模式不为'-'或'无操作'）。")
        if not messagebox.askyesno("确认操作", f"即将处理 {len(items_to_process)} 个商品。此操作不可逆，是否继续？"): return
        self.set_ui_state(True)
        threading.Thread(target=self._batch_process_thread, args=(items_to_process,), daemon=True).start()

    def start_auto_match_images(self):
        if not self.image_dir:
            messagebox.showerror("错误", "请先选择一个图片文件夹。")
            return
        
        add_items = []
        for item_id in self.product_tree.get_children():
            if self.product_tree.set(item_id, "action_mode") == "重创":
                add_items.append(self.product_tree.set(item_id, "excel_title"))

        if not add_items:
            messagebox.showinfo("提示", "表格中没有找到需要'重创'的新增套餐。")
            return

        self.set_ui_state(True)
        threading.Thread(target=self._auto_match_images_thread, args=(add_items,), daemon=True).start()

    def _auto_match_images_thread(self, add_items):
        self.log("--- 开始智能匹配套餐和图片 ---")
        if not llm_client:
            self.log("[Error] LLM客户端未初始化，无法进行智能匹配。")
            self.master.after(0, lambda: self.set_ui_state(False))
            return
        
        # 1. AI分析图片文件夹中的图片
        self.log("步骤 1/2: 正在使用AI分析图片内容...")
        image_summaries = []
        supported_formats = (".jpg", ".jpeg", ".png", ".bmp")
        model_index = 0
        try:
            image_files = [f for f in os.listdir(self.image_dir) if f.lower().endswith(supported_formats)]
            for filename in image_files:
                try:
                    full_path = os.path.join(self.image_dir, filename)
                    with open(full_path, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                    
                    image_url = f"data:image/jpeg;base64,{base64_image}"
                    
                    selected_vision_model = VISION_MODEL_IDS[model_index % len(VISION_MODEL_IDS)]
                    self.log(f"使用视觉模型: {selected_vision_model}")
                    model_index += 1
                    
                    response = llm_client.chat.completions.create(
                        model=selected_vision_model,
                        messages=[{
                            'role': 'user',
                            'content': [{
                                'type': 'text',
                                'text': '根据图片内容，为其生成一个简短且描述性的中文名（不要带扩展名）。例如，如果图片是关于500元网费套餐，就返回"500元网费"。',
                            }, {
                                'type': 'image_url',
                                'image_url': { 'url': image_url },
                            }],
                        }]
                    )
                    summary = response.choices[0].message.content.strip()
                    self.log(f"--- [Vision LLM Raw Response for Image Summary] ---\n{summary}\n" + "-"*30)
                    image_summaries.append({"original_filename": filename, "summary": summary})
                    self.log(f"分析图片 '{filename}' -> AI摘要: '{summary}'")
                except Exception as e:
                    self.log(f"[Error] 分析图片 {filename} 时出错: {e}")
        except Exception as e:
            self.log(f"[Error] 遍历图片文件夹时出错: {e}")
            self.master.after(0, lambda: self.set_ui_state(False))
            return

        if not image_summaries:
            self.log("[Error] 未能成功分析任何图片。")
            self.master.after(0, lambda: self.set_ui_state(False))
            return

        # 2. AI匹配文本
        self.log("步骤 2/2: 正在使用AI匹配图片摘要和套餐标题...")
        image_summary_list = [item['summary'] for item in image_summaries]
        
        multi_mode = self.multi_match_mode_var.get()
        if multi_mode:
            self.log("--- 当前为 [图片匹配多套餐] 模式 ---")
            prompt = f"""
            现有以下需要创建的套餐列表：
            {json.dumps(add_items, ensure_ascii=False)}

            以及以下从图片中分析出的摘要列表：
            {json.dumps(image_summary_list, ensure_ascii=False)}

            任务：对于每一个“图片摘要”，判断它可以匹配到哪些“套餐列表”中的项目。匹配应该是基于核心关键词的包含关系。
            例如，摘要“网费”可以匹配所有标题中包含“网费”的套餐。
            
            返回一个严格的JSON对象，其中键是图片摘要，值是一个包含所有匹配的套餐标题的**列表**。如果一个摘要找不到任何匹配项，值应为空列表 `[]`。
            例如: {{ "通用网费图": ["【新客】50元网费", "【老客】100元网费"], "包时套餐图": [] }}
            """
        else:
            self.log("--- 当前为 [一对一精准匹配] 模式 ---")
            prompt = f"""
            现有以下需要创建的套餐列表：
            {json.dumps(add_items, ensure_ascii=False)}

            以及以下从图片中分析出的摘要列表：
            {json.dumps(image_summary_list, ensure_ascii=False)}

            请为每个“套餐列表”中的项目，在“图片摘要列表”中找到**最匹配**的一项。
            返回一个严格的JSON对象，其中键是套餐标题，值是匹配上的图片摘要。如果找不到匹配项，请将值设为 null。
            例如: {{ "【新客】50元网费": "50元网费", "【专享】300元包时": null }}
            """

        try:
            response = llm_client.chat.completions.create(
                model=LLM_MODEL_ID, # 使用文本模型
                messages=[
                    {'role': 'system', 'content': 'You are a helpful assistant that only returns JSON.'},
                    {'role': 'user', 'content': prompt}
                ]
            )
            
            match_result_str = response.choices[0].message.content
            log_func(f"--- [LLM Raw Response for Image-Text Matching] ---\n{match_result_str}\n" + "-"*30)
            json_match = re.search(r'\{[\s\S]*\}', match_result_str)
            if json_match:
                cleaned_response = json_match.group(0)
                match_result = json.loads(cleaned_response)
            else:
                raise json.JSONDecodeError("在LLM响应中未找到JSON对象", match_result_str, 0)

            self.log("智能匹配API调用成功，正在更新UI...")

            summary_to_filename = {item['summary']: item['original_filename'] for item in image_summaries}

            # 在UI线程中更新Treeview
            def _update_ui():
                title_to_image_map = {}
                if multi_mode:
                    self.log("[Debug] 进入多对多匹配UI更新逻辑。")
                    # "一对多"逻辑: 反转字典
                    for summary, titles in match_result.items():
                        if summary in summary_to_filename:
                            filename = summary_to_filename[summary]
                            for title in titles:
                                title_to_image_map[title] = filename
                else:
                    self.log("[Debug] 进入一对一匹配UI更新逻辑。")
                    # "一对一"逻辑
                    for title, summary in match_result.items():
                        if summary in summary_to_filename:
                            title_to_image_map[title] = summary_to_filename[summary]
                
                self.log(f"[Debug] 最终构建的 '套餐标题 -> 图片' 映射: {json.dumps(title_to_image_map, ensure_ascii=False, indent=2)}")

                for item_id in self.product_tree.get_children():
                    if self.product_tree.set(item_id, "action_mode") == "重创":
                        title = self.product_tree.set(item_id, "excel_title")
                        if title in title_to_image_map:
                            filename = title_to_image_map[title]
                            self.product_tree.set(item_id, "matched_image", filename)
                            self.log(f"UI更新: 套餐 '{title}' -> 图片 '{filename}'")
                        else:
                            self.log(f"[Debug] 套餐 '{title}' 在映射中未找到匹配图片。")
                
                self.log("--- 智能匹配完成 ---")
                self.set_ui_state(False)

            self.master.after(0, _update_ui)

        except json.JSONDecodeError:
            self.log(f"[Error] 解析LLM返回的JSON失败。返回内容: {match_result_str}")
            self.master.after(0, lambda: messagebox.showerror("AI匹配错误", "AI服务返回了无效的数据格式。"))
            self.master.after(0, lambda: self.set_ui_state(False))
        except Exception as e:
            self.log(f"[Error] 智能匹配过程中出错: {e}")
            self.master.after(0, lambda: messagebox.showerror("AI匹配错误", f"请求AI服务时出错: {e}"))
            self.master.after(0, lambda: self.set_ui_state(False))

    def _batch_process_thread(self, items_to_process):
        success_count, failed_items = 0, []
        items_to_process.sort(key=lambda x: 1 if x["action_mode"] != "修改" else 0)
        
        # 为"重创"模式准备模板ID：
        # 1. 优先使用第一个"修改"操作的商品ID
        # 2. 否则使用当前门店商品列表中的第一个（self.douyin_products是当前门店的套餐列表）
        template_id_for_recreate = next(
            (item['product_id'] for item in items_to_process if item['action_mode'] == "修改"), 
            self.douyin_products[0]['id'] if self.douyin_products else None
        )
        
        if not template_id_for_recreate and any(item['action_mode'] == "重创" for item in items_to_process):
            self.log("[Error] 重创模式需要模板商品，但当前门店没有可用的商品。")
            self.master.after(0, lambda: messagebox.showerror("错误", "重创模式需要模板商品，但当前门店没有可用的商品。\n请先查询门店商品列表。"))
            self.master.after(0, lambda: self.set_ui_state(False))
            return

        # 记录模板商品信息（用于重创模式）
        if template_id_for_recreate and any(item['action_mode'] == "重创" for item in items_to_process):
            template_product = next((p for p in self.douyin_products if p['id'] == template_id_for_recreate), None)
            if template_product:
                self.log(f"--- 重创模式将使用模板商品: {template_product['name']} (ID: {template_id_for_recreate}) ---")
            else:
                self.log(f"--- 重创模式将使用模板商品ID: {template_id_for_recreate} ---")

        for item in items_to_process:
            mode = item["action_mode"]
            product_id = item["product_id"]
            if mode == "下架":
                success, reason = operate_douyin_product(self.douyin_access_token, product_id, self.log, offline=True)
            elif mode == "修改":
                # 修改模式：使用当前商品自己的ID作为模板
                success, reason = update_douyin_product(self.douyin_access_token, product_id, item["new_data"], self.log, mode, image_dir=self.image_dir, target_poi_id=self.current_poi_id)
            else:  # 重创模式 - 使用网页端API
                # 使用网页端API创建商品（复用模板图片，创建后自动修改POI ID）
                # 模板来源：当前门店的商品列表中的第一个
                product_id_created, reason = create_product_via_web(
                    DOUYIN_WEB_COOKIE,
                    DOUYIN_WEB_CSRF_TOKEN,
                    DOUYIN_ROOT_LIFE_ACCOUNT_ID,
                    template_id_for_recreate,  # 模板商品ID（来自当前门店）
                    item["new_data"],
                    self.current_poi_id,  # 目标门店POI ID
                    self.douyin_access_token,  # 用于后续修改POI ID
                    self.log
                )
                success = product_id_created is not None
            
            if success: success_count += 1
            else: failed_items.append(f"ID {product_id}: {reason}")
            time.sleep(1)

        summary_message = f"批量操作完成！\n\n成功: {success_count} 个\n失败: {len(failed_items)} 个"
        if failed_items:
            self.log("--- 操作失败详情 ---"); [self.log(f) for f in failed_items]
            summary_message += "\n\n失败详情请查看日志。"
        self.master.after(0, lambda: messagebox.showinfo("操作完成", summary_message))
        self.master.after(0, self.start_query_douyin)

if __name__ == "__main__":
    root = tk.Tk()
    app = App(master=root)
    root.mainloop()
